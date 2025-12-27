"""
Payment Summary (Royalty Report) Generator
Generates royalty calculation reports based on RVCR data

This module can be run standalone for testing:
    python -m services.payment_summary_generator

Or imported and used in routes:
    from services.payment_summary_generator import (
        calculate_payment_summary,
        generate_payment_summary_excel,
        convert_excel_to_pdf
    )

Royalty calculation rules based on SERVPRO Franchise License Agreement.
Reference: Royalty Calculation/Royalties.pdf
"""

import os
import json
import tempfile
import re
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================================
# ROYALTY CALCULATION CONSTANTS
# Based on SERVPRO Royalty Table (Royalties.pdf)
# ============================================================================

# Standard Rate Royalty Tiers (excluding Subcontract and Reconstruction)
# Monthly Gross Volume -> Royalty %, Fixed Royalty, Fixed Fee
STANDARD_RATE_TIERS = [
    {"min": 0, "max": 12849.99, "rate": 0.10, "fixed": 0, "fixed_fee": 45},
    {"min": 12850, "max": 21416.99, "rate": 0.09, "fixed": 0, "fixed_fee": 65},
    {"min": 21417, "max": 32124.99, "rate": 0.08, "fixed": 0, "fixed_fee": 85},
    {"min": 32125, "max": 42832.99, "rate": 0.075, "fixed": 0, "fixed_fee": 95},
    {"min": 42833, "max": 74956.99, "rate": 0.07, "fixed": 0, "fixed_fee": 115},
    {"min": 74957, "max": 107082.99, "rate": 0.065, "fixed": 5247.00, "threshold": 74956.99, "fixed_fee": 115},
    {"min": 107083, "max": 160623.99, "rate": 0.06, "fixed": 7335.00, "threshold": 107082.99, "fixed_fee": 115},
    {"min": 160624, "max": 214164.99, "rate": 0.055, "fixed": 10547.00, "threshold": 160623.99, "fixed_fee": 115},
    {"min": 214165, "max": float('inf'), "rate": 0.05, "fixed": 13492.00, "threshold": 214164.99, "fixed_fee": 115},
]

# Minimum Royalty when no revenue is collected
MINIMUM_ROYALTY = 145.00

# Reduced Rate Royalty (Subcontract and Reconstruction)
# Up to $32,124.99: 4%, above: 3% on excess + $1,285 fixed
REDUCED_RATE_THRESHOLD = 32124.99
REDUCED_RATE_BELOW = 0.04  # 4% for amounts up to threshold
REDUCED_RATE_ABOVE = 0.03  # 3% for amounts above threshold
REDUCED_RATE_FIXED = 1285.00  # Fixed amount when above threshold

# Fee Rates
# National Accounts Core/Regional Fee: 0.5% of gross volume (excluding Subcontract AND Reconstruction), no cap
NATIONAL_ACCOUNTS_FEE_RATE = 0.005  # 0.5%

# National Brand Fund Fee: 2.5% of gross volume (excluding Subcontract ONLY, includes Reconstruction), capped annually
# Per Royalties.pdf: "2.5 percent of reported gross volume exclusive of Subcontract"
NATIONAL_BRAND_FUND_RATE = 0.025  # 2.5%
NATIONAL_BRAND_FUND_CAP = {
    2017: 800000,
    2018: 850000,
    2019: 900000,
    2020: 950000,
    2021: 1050000,
    2022: 1150000,
    2023: 1250000,
    2024: 1350000,
    2025: 1450000,
}

# National Brand Fund Fee - Reduced Rate Services: 0.25% of Subcontract and Reconstruction, no cap
NATIONAL_BRAND_FUND_REDUCED_RATE = 0.0025  # 0.25%


# ============================================================================
# CATEGORY MAPPING FOR DYNAMIC COLUMN EXTRACTION
# ============================================================================

# Standard category names that appear in RVCR reports
CATEGORY_PATTERNS = {
    "water": [
        r"Total Commercial - Water",
        r"Total Residential - Water",
        r"Commercial - Water",
        r"Residential - Water",
        r"Total 1 - WATER",
        r"1 - WATER",
    ],
    "fire": [
        r"Total Commercial - Fire",
        r"Total Residential - Fire",
        r"Commercial - Fire",
        r"Residential - Fire",
        r"Total 2 - FIRE",
        r"2 - FIRE",
    ],
    "mold_bio": [
        r"Total Commercial - Mold/Bio Hazard",
        r"Total Residential - Mold/Bio Hazard",
        r"Commercial - Mold/Bio Hazard",
        r"Residential - Mold/Bio Hazard",
        r"Total 3 - MOLD/BIO HAZARD",
        r"3 - MOLD/BIO HAZARD",
    ],
    "other": [
        r"Total Commercial - Other",
        r"Total Residential - Other",
        r"Commercial - Other",
        r"Residential - Other",
        r"Total 4 - OTHER",
        r"4 - OTHER",
    ],
    "subcontract": [
        r"Total Commercial - Subcontract",
        r"Total Residential - Subcontract",
        r"Commercial - Subcontract",
        r"Residential - Subcontract",
        r"Total 5 - SUBCONTRACT",
        r"5 - SUBCONTRACT",
    ],
    "reconstruction": [
        r"Total Commercial - Reconstruction",
        r"Total Residential - Reconstruction",
        r"Commercial - Reconstruction",
        r"Residential - Reconstruction",
        r"Total 6 - RECONSTRUCTION",
        r"6 - RECONSTRUCTION",
    ],
}


# ============================================================================
# ROYALTY CALCULATION FUNCTIONS
# ============================================================================

def calculate_standard_rate_royalty(volume: float) -> dict:
    """
    Calculate royalty for Standard Rate Services using tiered table.
    
    Standard Rate Services include: Water, Fire, Mold/Bio Hazard, Other
    (excludes Subcontract and Reconstruction)
    
    Returns dict with royalty amount, tier info, and description.
    """
    if volume <= 0:
        return {
            "royalty": MINIMUM_ROYALTY,
            "fixed_fee": 45,
            "description": f"Minimum royalty (no revenue): ${MINIMUM_ROYALTY:.2f}",
            "rate": 0,
            "tier_description": "Minimum",
            "is_minimum": True
        }

    for tier in STANDARD_RATE_TIERS:
        if volume <= tier["max"]:
            if "threshold" in tier:
                # Tiered calculation: fixed + rate on excess
                excess = volume - tier["threshold"]
                royalty = tier["fixed"] + (excess * tier["rate"])
                description = f"${tier['fixed']:,.2f} + {tier['rate']*100:.1f}% of ${excess:,.2f}"
                tier_desc = f"${tier['min']:,.2f} - ${tier['max']:,.2f}"
            else:
                # Simple percentage
                royalty = volume * tier["rate"]
                description = f"{tier['rate']*100:.1f}% of ${volume:,.2f}"
                tier_desc = f"${tier['min']:,.2f} - ${tier['max']:,.2f}"

            return {
                "royalty": round(royalty, 2),
                "fixed_fee": tier["fixed_fee"],
                "description": description,
                "rate": tier["rate"],
                "tier_description": tier_desc,
                "is_minimum": False
            }

    # Should not reach here
    return {
        "royalty": 0,
        "fixed_fee": 115,
        "description": "Error calculating royalty",
        "rate": 0,
        "tier_description": "N/A",
        "is_minimum": False
    }


def calculate_reduced_rate_royalty(volume: float) -> dict:
    """
    Calculate royalty for Reduced Rate Services (Subcontract + Reconstruction).
    
    Calculation:
    - Up to $32,124.99: 4%
    - Above $32,124.99: $1,285.00 + 3% on excess
    
    Returns dict with royalty amount and description.
    """
    if volume <= 0:
        return {
            "royalty": 0,
            "description": "No reduced rate revenue",
            "excess_amount": 0,
            "rate_applied": 0
        }

    if volume <= REDUCED_RATE_THRESHOLD:
        royalty = volume * REDUCED_RATE_BELOW
        return {
            "royalty": round(royalty, 2),
            "description": f"{REDUCED_RATE_BELOW*100:.0f}% of ${volume:,.2f}",
            "excess_amount": 0,
            "rate_applied": REDUCED_RATE_BELOW
        }
    else:
        excess = volume - REDUCED_RATE_THRESHOLD
        royalty = REDUCED_RATE_FIXED + (excess * REDUCED_RATE_ABOVE)
        return {
            "royalty": round(royalty, 2),
            "description": f"${REDUCED_RATE_FIXED:,.2f} + {REDUCED_RATE_ABOVE*100:.0f}% of ${excess:,.2f}",
            "excess_amount": round(excess, 2),
            "rate_applied": REDUCED_RATE_ABOVE
        }


def get_national_brand_fund_cap(year: int) -> float:
    """Get the National Brand Fund cap for a given year."""
    if year in NATIONAL_BRAND_FUND_CAP:
        return NATIONAL_BRAND_FUND_CAP[year]
    # For future years, assume $100,000 increase per year from 2025
    if year > 2025:
        return NATIONAL_BRAND_FUND_CAP[2025] + (year - 2025) * 100000
    # For years before 2017, use 2017 cap
    return NATIONAL_BRAND_FUND_CAP[2017]


def calculate_national_brand_fund_fee(
    standard_rate_this_month: float,
    standard_rate_ytd_before: float,
    year: int = 2025
) -> dict:
    """
    Calculate National Brand Fund Fee with annual cap.
    
    2.5% of Standard Rate Services (excluding Subcontract and Reconstruction),
    capped at annual limit (e.g., $1,450,000 for 2025).
    
    The cap increases by $100,000 each year.
    """
    cap = get_national_brand_fund_cap(year)
    
    if standard_rate_this_month <= 0:
        return {
            "fee": 0,
            "description": "No standard rate revenue",
            "capped": False,
            "remaining_cap": max(0, cap - standard_rate_ytd_before),
            "annual_cap": cap
        }

    remaining_cap = max(0, cap - standard_rate_ytd_before)

    if remaining_cap <= 0:
        # Cap already reached
        return {
            "fee": 0,
            "description": "Annual cap reached",
            "capped": True,
            "remaining_cap": 0,
            "annual_cap": cap
        }

    # Calculate fee on the lesser of this month's amount or remaining cap
    applicable_amount = min(standard_rate_this_month, remaining_cap)
    fee = applicable_amount * NATIONAL_BRAND_FUND_RATE

    capped = applicable_amount < standard_rate_this_month

    return {
        "fee": round(fee, 2),
        "description": f"{NATIONAL_BRAND_FUND_RATE*100:.1f}% of ${applicable_amount:,.2f}",
        "capped": capped,
        "remaining_cap": round(remaining_cap - applicable_amount, 2) if not capped else 0,
        "annual_cap": cap
    }


# ============================================================================
# DYNAMIC DATA EXTRACTION FROM RVCR JSON
# ============================================================================

def extract_column_mapping(rvcr_data: dict) -> Dict[str, int]:
    """
    Dynamically extract column indices from RVCR JSON data.
    Returns a mapping of column title to index.
    """
    columns = rvcr_data.get("Columns", {}).get("Column", [])
    mapping = {}
    for i, col in enumerate(columns):
        title = col.get("ColTitle", "")
        if title:
            mapping[title] = i
    return mapping


def find_category_column(col_mapping: Dict[str, int], patterns: List[str], 
                         type_filter: str = None, prefer_total: bool = True) -> Tuple[Optional[int], Optional[str]]:
    """
    Find a column index matching one of the patterns.
    
    Args:
        col_mapping: Dict mapping column titles to indices
        patterns: List of patterns to search for
        type_filter: Optional filter for 'commercial', 'residential', or 'total'
        prefer_total: If True, prefer columns starting with "Total" over others
    
    Returns:
        Tuple of (column index, matched title) or (None, None) if not found
    """
    # Sort column titles to prioritize "Total" columns first
    sorted_items = sorted(col_mapping.items(), 
                          key=lambda x: (0 if x[0].startswith("Total ") else 1, x[1]))
    
    for title, idx in sorted_items:
        title_lower = title.lower()
        
        # Apply type filter if specified
        if type_filter:
            if type_filter == "commercial" and "commercial" not in title_lower:
                continue
            if type_filter == "residential" and "residential" not in title_lower:
                continue
            if type_filter == "total" and ("commercial" in title_lower or "residential" in title_lower):
                continue
        
        # If prefer_total, skip non-Total columns for Commercial/Residential
        if prefer_total and type_filter in ["commercial", "residential"]:
            if not title.startswith("Total "):
                continue
        
        for pattern in patterns:
            if re.match(pattern, title, re.IGNORECASE) or title == pattern:
                return idx, title
    
    # If prefer_total didn't find anything, try again without the preference
    if prefer_total:
        return find_category_column(col_mapping, patterns, type_filter, prefer_total=False)
    
    return None, None


def extract_category_totals(rvcr_data: dict, verbose: bool = False) -> dict:
    """
    Dynamically extract category totals from RVCR JSON data.
    Returns dict with all category values for Commercial/Residential.
    """
    results = {
        "water": {"commercial": 0, "residential": 0, "total": 0},
        "fire": {"commercial": 0, "residential": 0, "total": 0},
        "mold_bio": {"commercial": 0, "residential": 0, "total": 0},
        "other": {"commercial": 0, "residential": 0, "total": 0},
        "subcontract": {"commercial": 0, "residential": 0, "total": 0},
        "reconstruction": {"commercial": 0, "residential": 0, "total": 0},
    }

    # Build column mapping
    col_mapping = extract_column_mapping(rvcr_data)
    if verbose:
        print(f"Found {len(col_mapping)} columns in RVCR data")

    # Find TOTAL row (GrandTotal section)
    rows = rvcr_data.get("Rows", {}).get("Row", [])
    total_row = None
    
    for row in rows:
        if row.get("type") == "Section" and row.get("group") == "GrandTotal":
            summary = row.get("Summary", {})
            total_row = summary.get("ColData", [])
            break
    
    if not total_row:
        if verbose:
            print("Could not find GrandTotal row in RVCR data")
        return results

    if verbose:
        print(f"Found TOTAL row with {len(total_row)} values")

    # Extract values for each category
    for category, patterns in CATEGORY_PATTERNS.items():
        # Try to find Commercial column
        commercial_patterns = [p for p in patterns if "Commercial" in p]
        idx, title = find_category_column(col_mapping, commercial_patterns)
        if idx is not None and idx < len(total_row):
            value_str = total_row[idx].get("value", "0")
            try:
                results[category]["commercial"] = float(value_str) if value_str else 0
                if verbose:
                    print(f"  {title} (idx {idx}): {results[category]['commercial']:.2f}")
            except ValueError:
                pass
        
        # Try to find Residential column
        residential_patterns = [p for p in patterns if "Residential" in p]
        idx, title = find_category_column(col_mapping, residential_patterns)
        if idx is not None and idx < len(total_row):
            value_str = total_row[idx].get("value", "0")
            try:
                results[category]["residential"] = float(value_str) if value_str else 0
                if verbose:
                    print(f"  {title} (idx {idx}): {results[category]['residential']:.2f}")
            except ValueError:
                pass
        
        # Try to find Total column (fallback)
        total_patterns = [p for p in patterns if p.startswith("Total ") and "Commercial" not in p and "Residential" not in p]
        idx, title = find_category_column(col_mapping, total_patterns)
        if idx is not None and idx < len(total_row):
            value_str = total_row[idx].get("value", "0")
            try:
                results[category]["total"] = float(value_str) if value_str else 0
                if verbose:
                    print(f"  {title} (idx {idx}): {results[category]['total']:.2f}")
            except ValueError:
                pass
        
        # Calculate total from commercial + residential if we have both
        if results[category]["commercial"] > 0 or results[category]["residential"] > 0:
            results[category]["total"] = results[category]["commercial"] + results[category]["residential"]

    # Log extracted totals
    if verbose:
        print("\nExtracted category totals:")
        for cat_name, cat_data in results.items():
            if cat_data["total"] > 0:
                print(f"  {cat_name}: Commercial=${cat_data['commercial']:.2f}, "
                      f"Residential=${cat_data['residential']:.2f}, Total=${cat_data['total']:.2f}")

    return results


def extract_period_info(rvcr_data: dict) -> Tuple[Optional[int], Optional[int], Optional[str], Optional[str]]:
    """
    Extract period information from RVCR JSON header.
    Returns: (month, year, start_date, end_date)
    """
    header = rvcr_data.get("Header", {})
    start_period = header.get("StartPeriod", "")
    end_period = header.get("EndPeriod", "")
    
    month = None
    year = None
    
    if end_period:
        try:
            # Parse YYYY-MM-DD format
            parts = end_period.split("-")
            if len(parts) >= 2:
                year = int(parts[0])
                month = int(parts[1])
        except (ValueError, IndexError):
            pass
    
    return month, year, start_period, end_period


# ============================================================================
# PAYMENT SUMMARY CALCULATION
# ============================================================================

def calculate_payment_summary(
    last_month_data: dict,
    ytd_data: dict,
    franchise_number: str,
    department_name: str,
    owner_name: str = "",
    period_month: int = None,
    period_year: int = None,
    verbose: bool = False
) -> dict:
    """
    Calculate complete payment summary from RVCR data.
    
    Based on SERVPRO Royalty guidelines (Royalties.pdf):
    - Standard Rate Services: Water, Fire, Mold/Bio Hazard, Other
    - Reduced Rate Services: Subcontract, Reconstruction
    - Fixed Fee based on volume tier
    - National Accounts Fee: 0.5% of Standard Rate (no cap)
    - National Brand Fund Fee: 2.5% of Standard Rate (capped annually)
    - National Brand Fund Fee Reduced: 0.25% of Reduced Rate (no cap)
    """
    # Extract period info if not provided
    if period_month is None or period_year is None:
        extracted_month, extracted_year, _, _ = extract_period_info(last_month_data)
        if period_month is None:
            period_month = extracted_month or datetime.now().month
        if period_year is None:
            period_year = extracted_year or datetime.now().year
    
    # Extract category totals for this month
    if verbose:
        print("\n=== Extracting Last Month Data ===")
    categories = extract_category_totals(last_month_data, verbose=verbose)

    # Extract YTD totals
    if verbose:
        print("\n=== Extracting YTD Data ===")
    ytd_categories = extract_category_totals(ytd_data, verbose=verbose)

    # Calculate category totals - This Month
    water_total = categories["water"]["total"]
    fire_total = categories["fire"]["total"]
    mold_total = categories["mold_bio"]["total"]
    other_total = categories["other"]["total"]
    subcontract_total = categories["subcontract"]["total"]
    reconstruction_total = categories["reconstruction"]["total"]

    standard_rate_total = water_total + fire_total + mold_total + other_total
    reduced_rate_total = subcontract_total + reconstruction_total
    grand_total = standard_rate_total + reduced_rate_total

    # Calculate subtotals - YTD
    ytd_water = ytd_categories["water"]["total"]
    ytd_fire = ytd_categories["fire"]["total"]
    ytd_mold = ytd_categories["mold_bio"]["total"]
    ytd_other = ytd_categories["other"]["total"]
    ytd_subcontract = ytd_categories["subcontract"]["total"]
    ytd_reconstruction = ytd_categories["reconstruction"]["total"]

    ytd_standard_rate = ytd_water + ytd_fire + ytd_mold + ytd_other
    ytd_reduced_rate = ytd_subcontract + ytd_reconstruction
    ytd_grand_total = ytd_standard_rate + ytd_reduced_rate

    # Calculate YTD before this month (for cap calculations)
    ytd_standard_before = max(0, ytd_standard_rate - standard_rate_total)

    # Calculate royalties for Standard Rate Services
    standard_royalty = calculate_standard_rate_royalty(standard_rate_total)

    # Calculate royalties for Reduced Rate Services
    reduced_royalty = calculate_reduced_rate_royalty(reduced_rate_total)

    # Calculate proportional royalty rate for allocation to subcategories
    std_royalty_rate = standard_royalty["royalty"] / standard_rate_total if standard_rate_total > 0 else 0
    red_royalty_rate = reduced_royalty["royalty"] / reduced_rate_total if reduced_rate_total > 0 else 0

    # Calculate royalties for each subcategory (Commercial and Residential separately)
    def calc_subcategory_royalties(cat_data: dict, rate: float) -> dict:
        return {
            "commercial": round(cat_data["commercial"] * rate, 2),
            "residential": round(cat_data["residential"] * rate, 2),
            "total": round(cat_data["total"] * rate, 2)
        }

    water_royalties = calc_subcategory_royalties(categories["water"], std_royalty_rate)
    fire_royalties = calc_subcategory_royalties(categories["fire"], std_royalty_rate)
    mold_royalties = calc_subcategory_royalties(categories["mold_bio"], std_royalty_rate)
    other_royalties = calc_subcategory_royalties(categories["other"], std_royalty_rate)
    subcontract_royalties = calc_subcategory_royalties(categories["subcontract"], red_royalty_rate)
    reconstruction_royalties = calc_subcategory_royalties(categories["reconstruction"], red_royalty_rate)

    # Calculate fees
    fixed_fee = standard_royalty["fixed_fee"]
    
    # National Accounts Fee: 0.5% of Standard Rate (excluding Subcontract AND Reconstruction)
    national_accounts_fee = round(standard_rate_total * NATIONAL_ACCOUNTS_FEE_RATE, 2)

    # National Brand Fund Fee: 2.5% of Standard Rate with annual cap
    # Cap is based on YTD Standard Rate revenue (first $1,450,000 in 2025)
    # Note: Despite PDF saying "exclusive of Subcontract", the SERVPRO system
    # applies this fee to Standard Rate only (Water, Fire, Mold, Other)
    # Reconstruction is covered by the 0.25% Reduced Rate fee instead
    brand_fund = calculate_national_brand_fund_fee(
        standard_rate_total,
        ytd_standard_before,
        year=period_year
    )
    national_brand_fee = brand_fund["fee"]

    # National Brand Fund Fee - Reduced Rate: 0.25% of Subcontract and Reconstruction (no cap)
    national_brand_reduced_fee = round(reduced_rate_total * NATIONAL_BRAND_FUND_REDUCED_RATE, 2)

    # Total payable
    total_royalty_payable = standard_royalty["royalty"] + reduced_royalty["royalty"]

    total_fees_payable = (
        fixed_fee +
        national_accounts_fee +
        national_brand_fee +
        national_brand_reduced_fee
    )

    grand_total_payable = total_royalty_payable + total_fees_payable

    return {
        "franchise_number": franchise_number,
        "department_name": department_name,
        "owner_name": owner_name,
        "period_month": period_month,
        "period_year": period_year,
        "categories": {
            "water": {
                "commercial": {"this_month": categories["water"]["commercial"], 
                              "ytd": ytd_categories["water"]["commercial"], 
                              "royalty": water_royalties["commercial"]},
                "residential": {"this_month": categories["water"]["residential"], 
                               "ytd": ytd_categories["water"]["residential"], 
                               "royalty": water_royalties["residential"]},
                "total": {"this_month": water_total, "ytd": ytd_water, "royalty": water_royalties["total"]}
            },
            "fire": {
                "commercial": {"this_month": categories["fire"]["commercial"], 
                              "ytd": ytd_categories["fire"]["commercial"], 
                              "royalty": fire_royalties["commercial"]},
                "residential": {"this_month": categories["fire"]["residential"], 
                               "ytd": ytd_categories["fire"]["residential"], 
                               "royalty": fire_royalties["residential"]},
                "total": {"this_month": fire_total, "ytd": ytd_fire, "royalty": fire_royalties["total"]}
            },
            "mold_bio": {
                "commercial": {"this_month": categories["mold_bio"]["commercial"], 
                              "ytd": ytd_categories["mold_bio"]["commercial"], 
                              "royalty": mold_royalties["commercial"]},
                "residential": {"this_month": categories["mold_bio"]["residential"], 
                               "ytd": ytd_categories["mold_bio"]["residential"], 
                               "royalty": mold_royalties["residential"]},
                "total": {"this_month": mold_total, "ytd": ytd_mold, "royalty": mold_royalties["total"]}
            },
            "other": {
                "commercial": {"this_month": categories["other"]["commercial"], 
                              "ytd": ytd_categories["other"]["commercial"], 
                              "royalty": other_royalties["commercial"]},
                "residential": {"this_month": categories["other"]["residential"], 
                               "ytd": ytd_categories["other"]["residential"], 
                               "royalty": other_royalties["residential"]},
                "total": {"this_month": other_total, "ytd": ytd_other, "royalty": other_royalties["total"]}
            },
            "subcontract": {
                "commercial": {"this_month": categories["subcontract"]["commercial"], 
                              "ytd": ytd_categories["subcontract"]["commercial"], 
                              "royalty": subcontract_royalties["commercial"]},
                "residential": {"this_month": categories["subcontract"]["residential"], 
                               "ytd": ytd_categories["subcontract"]["residential"], 
                               "royalty": subcontract_royalties["residential"]},
                "total": {"this_month": subcontract_total, "ytd": ytd_subcontract, "royalty": subcontract_royalties["total"]}
            },
            "reconstruction": {
                "commercial": {"this_month": categories["reconstruction"]["commercial"], 
                              "ytd": ytd_categories["reconstruction"]["commercial"], 
                              "royalty": reconstruction_royalties["commercial"]},
                "residential": {"this_month": categories["reconstruction"]["residential"], 
                               "ytd": ytd_categories["reconstruction"]["residential"], 
                               "royalty": reconstruction_royalties["residential"]},
                "total": {"this_month": reconstruction_total, "ytd": ytd_reconstruction, "royalty": reconstruction_royalties["total"]}
            },
        },
        "subtotals": {
            "standard_rate": {"this_month": standard_rate_total, "ytd": ytd_standard_rate},
            "reduced_rate": {"this_month": reduced_rate_total, "ytd": ytd_reduced_rate},
            "total": {"this_month": grand_total, "ytd": ytd_grand_total},
        },
        "royalties": {
            "standard_rate": {
                "amount": standard_royalty["royalty"],
                "description": standard_royalty["description"],
                "tier": standard_royalty["tier_description"],
                "is_minimum": standard_royalty.get("is_minimum", False),
            },
            "reduced_rate": {
                "amount": reduced_royalty["royalty"],
                "description": reduced_royalty["description"],
                "excess_amount": reduced_royalty["excess_amount"],
            },
            "total": total_royalty_payable,
        },
        "fees": {
            "fixed_fee": fixed_fee,
            "national_accounts": national_accounts_fee,
            "national_brand_fund": {
                "amount": national_brand_fee,
                "capped": brand_fund["capped"],
                "remaining_cap": brand_fund["remaining_cap"],
                "annual_cap": brand_fund["annual_cap"],
            },
            "national_brand_reduced": national_brand_reduced_fee,
            "total": total_fees_payable,
        },
        "grand_total_payable": grand_total_payable,
    }


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def generate_payment_summary_excel(summary: dict, output_path: str) -> str:
    """
    Generate Payment Summary Excel file matching the SERVPRO format.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Summary Report"

    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    normal_font = Font(size=10)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    subtotal_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    row = 1

    # Header section - "Royalty Reporting" to match SERVPRO format
    ws.merge_cells('A1:E1')
    header_cell = ws.cell(row=row, column=1, value="Royalty Reporting")
    header_cell.font = header_font
    header_cell.alignment = Alignment(horizontal='center')
    row += 2

    # Franchise info - matching SERVPRO format
    ws.cell(row=row, column=1, value="Franchise:")
    ws.cell(row=row, column=2, value=f"{summary['franchise_number']}   {summary['department_name']}")
    ws.cell(row=row, column=4, value="Owner:")
    ws.cell(row=row, column=5, value=summary.get("owner_name", ""))
    row += 1

    month_name = datetime(2000, summary["period_month"], 1).strftime("%m") if summary["period_month"] else "N/A"
    ws.cell(row=row, column=1, value="Month/Year:")
    ws.cell(row=row, column=2, value=f"{month_name}/{summary['period_year']}")
    ws.cell(row=row, column=4, value="Date of Mailing:")
    ws.cell(row=row, column=5, value=datetime.now().strftime("%m/%d/%Y"))
    row += 2

    # Column headers
    headers = ["Category", "", "Revenue This Month", "Year To Date", "Royalty Payable"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Category data - matching SERVPRO format
    categories_config = [
        ("Water Restoration", "water", "standard"),
        ("Fire Restoration", "fire", "standard"),
        ("Mold/Bio Haz. Restoration", "mold_bio", "standard"),
        ("Other", "other", "standard"),
        ("Subcontract (Mitigation)", "subcontract", "reduced"),
        ("Reconstruction (In-House & Subcontract)", "reconstruction", "reduced"),
    ]

    for cat_name, cat_key, rate_type in categories_config:
        cat_data = summary["categories"][cat_key]

        # Category header
        ws.cell(row=row, column=1, value=cat_name).font = section_font
        row += 1

        # Commercial with Royalty
        ws.cell(row=row, column=2, value="Commercial:")
        cell_rev = ws.cell(row=row, column=3, value=cat_data["commercial"]["this_month"])
        cell_rev.number_format = '#,##0.00'
        cell_rev.fill = yellow_fill
        cell_ytd = ws.cell(row=row, column=4, value=cat_data["commercial"]["ytd"])
        cell_ytd.number_format = '$#,##0.00'
        cell_royalty = ws.cell(row=row, column=5, value=cat_data["commercial"]["royalty"])
        cell_royalty.number_format = '$#,##0.00'
        row += 1

        # Residential with Royalty
        ws.cell(row=row, column=2, value="Residential:")
        cell_rev = ws.cell(row=row, column=3, value=cat_data["residential"]["this_month"])
        cell_rev.number_format = '#,##0.00'
        cell_rev.fill = yellow_fill
        cell_ytd = ws.cell(row=row, column=4, value=cat_data["residential"]["ytd"])
        cell_ytd.number_format = '$#,##0.00'
        cell_royalty = ws.cell(row=row, column=5, value=cat_data["residential"]["royalty"])
        cell_royalty.number_format = '$#,##0.00'
        row += 1

    # Subtotals - matching SERVPRO format
    subtotals = summary["subtotals"]
    royalties = summary["royalties"]
    fees = summary["fees"]

    # Standard Rate Subtotal
    cell = ws.cell(row=row, column=1, value="Subtotal: Standard Rate Services (Water, Fire, Mold/Bio Haz, Other)")
    cell.font = section_font
    cell = ws.cell(row=row, column=3, value=subtotals["standard_rate"]["this_month"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=4, value=subtotals["standard_rate"]["ytd"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=5, value=royalties["standard_rate"]["amount"])
    cell.number_format = '$#,##0.00'
    row += 1

    # Reduced Rate Subtotal
    cell = ws.cell(row=row, column=1, value="Subtotal: Reduced Rate Services (Subcontract, Reconstruction)")
    cell.font = section_font
    cell = ws.cell(row=row, column=3, value=subtotals["reduced_rate"]["this_month"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=4, value=subtotals["reduced_rate"]["ytd"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=5, value=royalties["reduced_rate"]["amount"])
    cell.number_format = '$#,##0.00'
    row += 1

    # Total
    cell = ws.cell(row=row, column=1, value="Total")
    cell.font = section_font
    cell = ws.cell(row=row, column=3, value=subtotals["total"]["this_month"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    cell = ws.cell(row=row, column=4, value=subtotals["total"]["ytd"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    cell = ws.cell(row=row, column=5, value=royalties["total"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    row += 1

    # Fixed Fee
    ws.cell(row=row, column=1, value="Fixed Fee")
    cell = ws.cell(row=row, column=5, value=fees["fixed_fee"])
    cell.number_format = '$#,##0.00'
    row += 1

    # National Accounts Fee
    ws.cell(row=row, column=1, value="National Accounts Core/Regional Fee (0.5%)")
    cell = ws.cell(row=row, column=5, value=fees["national_accounts"])
    cell.number_format = '$#,##0.00'
    row += 1

    # National Brand Fund Fee
    ws.cell(row=row, column=1, value="National Brand Fund Fee (2.5%)")
    cell = ws.cell(row=row, column=5, value=fees["national_brand_fund"]["amount"])
    cell.number_format = '$#,##0.00'
    row += 1

    # National Brand Fund Fee - Reduced Rate
    ws.cell(row=row, column=1, value="National Brand Fund Fee - Reduced Rate Services (0.25%)")
    cell = ws.cell(row=row, column=5, value=fees["national_brand_reduced"])
    cell.number_format = '$#,##0.00'
    row += 1

    # Royalty % And Fee Used section
    std_royalty = royalties["standard_rate"]
    red_royalty = royalties["reduced_rate"]

    ws.cell(row=row, column=1, value="Royalty % And Fee Used").font = section_font
    ws.cell(row=row, column=4, value="Total Royalty, Fixed Fee,").font = section_font
    cell = ws.cell(row=row, column=5, value=summary["grand_total_payable"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    row += 1

    # Standard rate description
    if std_royalty.get("is_minimum"):
        ws.cell(row=row, column=1, value=f"Minimum royalty: ${MINIMUM_ROYALTY:.2f}")
    else:
        ws.cell(row=row, column=1, value=f"pay {std_royalty['description']}")
    ws.cell(row=row, column=4, value="and Brand Fund Fee Payable").font = section_font
    row += 1

    ws.cell(row=row, column=1, value=f"plus ${fees['fixed_fee']:.2f} Fixed Fee")
    row += 1

    ws.cell(row=row, column=1, value=f"Reduced Rate Services pay {red_royalty['description']}")
    row += 2

    # Payment fields
    ws.cell(row=row, column=4, value="Check Number:")
    row += 1
    ws.cell(row=row, column=4, value="Check Amount:")

    # Save
    wb.save(output_path)
    return output_path


def convert_excel_to_pdf(excel_path: str, pdf_path: str) -> bool:
    """
    Attempt to convert Excel to PDF using available methods.
    Returns True if successful, False otherwise.
    """
    # Try LibreOffice first
    try:
        import subprocess
        output_dir = os.path.dirname(pdf_path)
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', output_dir, excel_path],
            capture_output=True,
            timeout=60
        )
        if result.returncode == 0:
            expected_pdf = os.path.join(output_dir, os.path.basename(excel_path).replace('.xlsx', '.pdf'))
            if os.path.exists(expected_pdf) and expected_pdf != pdf_path:
                os.rename(expected_pdf, pdf_path)
            return os.path.exists(pdf_path)
    except Exception as e:
        print(f"LibreOffice conversion failed: {e}")

    # Try win32com on Windows
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
        return os.path.exists(pdf_path)
    except Exception as e:
        print(f"win32com conversion failed: {e}")

    return False


# ============================================================================
# MAIN FUNCTION FOR TESTING
# ============================================================================

def main():
    """
    Main function to demonstrate usage.
    Similar to generate_royalty_report.py main function.
    """
    
    # Define file paths
    last_month_file = r"Royalty Calculation\rvcrsep.json"  # Result from Last Month API call
    ytd_file = r"Royalty Calculation\rvcrsepytd.json"  # Result from YTD API call
    output_file = r"Royalty Calculation\Generated_Payment_Summary_Test_3.xlsx"
    
    # Report configuration
    franchise_number = "11533"
    department_name = "Sarpy County"
    owner_name = "SERVPRO of Sooland, Lincoln, Sarpy County, Kearney & North Platte"
    
    print("="*70)
    print("PAYMENT SUMMARY GENERATOR")
    print("Based on SERVPRO Royalty Guidelines (Royalties.pdf)")
    print("="*70)
    
    # Load JSON files
    print(f"\nLoading last month data from: {last_month_file}")
    try:
        with open(last_month_file, 'r') as f:
            last_month_data = json.load(f)
        print("  Last month data loaded successfully")
    except FileNotFoundError:
        print(f"  ERROR: File not found: {last_month_file}")
        return
    except json.JSONDecodeError as e:
        print(f"  ERROR: Invalid JSON: {e}")
        return
    
    print(f"\nLoading YTD data from: {ytd_file}")
    try:
        with open(ytd_file, 'r') as f:
            ytd_data = json.load(f)
        print("  YTD data loaded successfully")
    except FileNotFoundError:
        print(f"  ERROR: File not found: {ytd_file}")
        return
    except json.JSONDecodeError as e:
        print(f"  ERROR: Invalid JSON: {e}")
        return
    
    # Extract period info from the data
    period_month, period_year, start_date, end_date = extract_period_info(last_month_data)
    print(f"\nPeriod detected: {start_date} to {end_date}")
    print(f"Month/Year: {period_month}/{period_year}")
    
    # Calculate payment summary
    print("\n" + "="*70)
    print("CALCULATING PAYMENT SUMMARY")
    print("="*70)
    
    summary = calculate_payment_summary(
        last_month_data=last_month_data,
        ytd_data=ytd_data,
        franchise_number=franchise_number,
        department_name=department_name,
        owner_name=owner_name,
        period_month=period_month,
        period_year=period_year,
        verbose=True
    )
    
    # Print results
    print("\n" + "="*70)
    print("PAYMENT SUMMARY RESULTS")
    print("="*70)
    print(f"\nFranchise: {summary['franchise_number']} - {summary['department_name']}")
    print(f"Owner: {summary['owner_name']}")
    print(f"Period: {summary['period_month']:02d}/{summary['period_year']}")
    
    print("\n--- Revenue Summary ---")
    print(f"Standard Rate Services (Water, Fire, Mold, Other):")
    print(f"  This Month: ${summary['subtotals']['standard_rate']['this_month']:,.2f}")
    print(f"  YTD:        ${summary['subtotals']['standard_rate']['ytd']:,.2f}")
    print(f"Reduced Rate Services (Subcontract, Reconstruction):")
    print(f"  This Month: ${summary['subtotals']['reduced_rate']['this_month']:,.2f}")
    print(f"  YTD:        ${summary['subtotals']['reduced_rate']['ytd']:,.2f}")
    print(f"Total Revenue:")
    print(f"  This Month: ${summary['subtotals']['total']['this_month']:,.2f}")
    print(f"  YTD:        ${summary['subtotals']['total']['ytd']:,.2f}")
    
    print("\n--- Royalties ---")
    std = summary['royalties']['standard_rate']
    red = summary['royalties']['reduced_rate']
    print(f"Standard Rate Royalty: ${std['amount']:,.2f}")
    print(f"  Calculation: {std['description']}")
    print(f"  Tier: {std['tier']}")
    if std.get('is_minimum'):
        print(f"  Note: Minimum royalty applied (${MINIMUM_ROYALTY:.2f})")
    print(f"Reduced Rate Royalty:  ${red['amount']:,.2f}")
    print(f"  Calculation: {red['description']}")
    print(f"Total Royalty:         ${summary['royalties']['total']:,.2f}")
    
    print("\n--- Fees ---")
    fees = summary['fees']
    print(f"Fixed Fee:                ${fees['fixed_fee']:,.2f}")
    print(f"National Accounts Fee:    ${fees['national_accounts']:,.2f} (0.5% of Standard Rate)")
    print(f"National Brand Fund Fee:  ${fees['national_brand_fund']['amount']:,.2f} (2.5% of Standard Rate)")
    if fees['national_brand_fund']['capped']:
        print(f"  Note: Annual cap reached (${fees['national_brand_fund']['annual_cap']:,.0f})")
    print(f"Brand Fund (Reduced):     ${fees['national_brand_reduced']:,.2f} (0.25% of Reduced Rate)")
    print(f"Total Fees:               ${fees['total']:,.2f}")
    
    print("\n" + "="*70)
    print(f"GRAND TOTAL PAYABLE: ${summary['grand_total_payable']:,.2f}")
    print("="*70)
    
    # Generate Excel file
    print(f"\nGenerating Excel report: {output_file}")
    try:
        generate_payment_summary_excel(summary, output_file)
        print("  Excel file generated successfully!")
    except Exception as e:
        print(f"  ERROR generating Excel: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Try PDF conversion
    pdf_path = output_file.replace('.xlsx', '.pdf')
    print(f"\nAttempting PDF conversion: {pdf_path}")
    if convert_excel_to_pdf(output_file, pdf_path):
        print("  PDF file generated successfully!")
    else:
        print("  PDF conversion not available (LibreOffice or Excel required)")
    
    print("\n" + "="*70)
    print("COMPLETE")
    print("="*70)
    
    return summary


if __name__ == "__main__":
    main()
