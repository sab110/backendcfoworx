"""
Dynamic Royalty Volume Calculation Report Generator
This script replicates Excel reports from QuickBooks JSON data files EXACTLY
"""

import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import os


class RoyaltyReportGenerator:
    """
    Generate royalty volume calculation reports from QuickBooks data
    Replicates the exact format of the original Excel template
    """
    
    def __init__(self):
        # Define styling constants to match original
        self.TITLE_FONT = Font(name='Arial', size=14, bold=True)
        self.SUBTITLE_FONT = Font(name='Arial', size=14, bold=True)
        self.DATE_FONT = Font(name='Arial', size=10, bold=True)
        self.HEADER_FONT = Font(name='Arial', size=9, bold=True)
        self.CATEGORY_FONT = Font(name='Arial', size=8, bold=True)
        self.DATA_FONT = Font(name='Arial', size=8, bold=False)
        self.TOTAL_FONT = Font(name='Arial', size=8, bold=True)
        
        self.CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
        self.LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
        self.RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    
    def load_json_file(self, filepath: str) -> Dict:
        """Load JSON file and return parsed data"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            raise Exception(f"Error loading {filepath}: {str(e)}")
    
    def format_currency(self, value: str) -> float:
        """Convert string currency to float"""
        try:
            if not value or value == '':
                return 0.0
            return float(value.replace(',', ''))
        except (ValueError, AttributeError):
            return 0.0
    
    def validate_report_data(self, json_data: Dict) -> bool:
        """
        Validate if the report contains data based on QB structure
        Checks the NoReportData option in Header
        """
        header = json_data.get('Header', {})
        options = header.get('Option', [])
        
        for option in options:
            if option.get('Name') == 'NoReportData':
                no_data = option.get('Value', 'false')
                # If NoReportData is "true", report has no data
                if no_data.lower() == 'true':
                    return False
        
        return True
    
    def get_report_metadata(self, json_data: Dict) -> Dict:
        """Extract metadata from report header"""
        header = json_data.get('Header', {})
        return {
            'report_name': header.get('ReportName', ''),
            'report_basis': header.get('ReportBasis', 'Cash'),
            'start_period': header.get('StartPeriod', ''),
            'end_period': header.get('EndPeriod', ''),
            'currency': header.get('Currency', 'USD'),
            'time': header.get('Time', ''),
            'date_macro': header.get('DateMacro', ''),
            'department': header.get('Department', ''),
            'class': header.get('Class', '')
        }
    
    def extract_totals_from_json(self, json_data: Dict) -> Dict:
        """
        Extract category totals from QuickBooks JSON
        Follows the official QuickBooks API structure:
        - Rows can be type="Section" (with Header, nested Rows, Summary) or type="Data" (just ColData)
        - Summary rows contain totals for sections
        - Uses full path keys to differentiate items with same name under different parents
        - Also captures Header-level values (unclassified amounts at subcategory level)
        """
        totals = {}
        
        def process_row(row, parent_path=""):
            """Recursively process rows according to QB structure"""
            
            row_type = row.get('type', '')
            current_name = ""
            
            # Handle Section rows (have Header, nested Rows, and Summary)
            if row_type == 'Section':
                # Get section name and Header-level value
                header_value = 0.0
                if 'Header' in row:
                    header_col_data = row['Header'].get('ColData', [])
                    if header_col_data:
                        current_name = header_col_data[0].get('value', '').strip()
                        # Get Header-level value (TOTAL column - last column)
                        header_total = header_col_data[-1].get('value', '0.00')
                        header_value = self.format_currency(header_total)
                
                # Build current path
                current_path = f"{parent_path}::{current_name}" if parent_path else current_name
                
                # Store Header-level value if non-zero (represents unclassified amounts)
                # Only store for subcategories (level 2+), not main categories
                if header_value != 0 and parent_path:
                    # Store as a special "__UNCLASSIFIED__" entry for this subcategory
                    totals[f"HEADER::{current_path}"] = header_value
                
                # Process nested rows with updated path
                if 'Rows' in row and 'Row' in row['Rows']:
                    for nested_row in row['Rows']['Row']:
                        process_row(nested_row, current_path)
                
                # Process Summary (contains the total for this section)
                if 'Summary' in row:
                    summary_col_data = row['Summary'].get('ColData', [])
                    if summary_col_data:
                        category = summary_col_data[0].get('value', '').strip()
                        total_value = summary_col_data[-1].get('value', '0.00')
                        
                        if category:
                            totals[category] = self.format_currency(total_value)
            
            # Handle Data rows (just have ColData directly)
            elif row_type == 'Data' or ('ColData' in row and 'Summary' not in row and 'Header' not in row):
                col_data = row.get('ColData', [])
                if col_data:
                    item_name = col_data[0].get('value', '').strip()
                    item_value = col_data[-1].get('value', '0.00')
                    
                    # Store data row values with FULL PATH to differentiate same-named items
                    if item_name and item_name not in ['Interest & Credit Card Fees', 'SD - Excise Tax']:
                        # Create full path key: parent_path::item_name
                        full_key = f"{parent_path}::{item_name}" if parent_path else item_name
                        totals[f"DATA::{full_key}"] = self.format_currency(item_value)
            
            # Handle rows without explicit type (legacy format or special cases)
            else:
                # Check for Summary (totals)
                if 'Summary' in row:
                    summary_col_data = row['Summary'].get('ColData', [])
                    if summary_col_data:
                        category = summary_col_data[0].get('value', '').strip()
                        total_value = summary_col_data[-1].get('value', '0.00')
                        
                        if category:
                            totals[category] = self.format_currency(total_value)
                
                # Process nested rows
                if 'Rows' in row and 'Row' in row['Rows']:
                    for nested_row in row['Rows']['Row']:
                        process_row(nested_row, parent_path)
        
        # Process top-level rows
        if 'Rows' in json_data and 'Row' in json_data['Rows']:
            for row in json_data['Rows']['Row']:
                process_row(row)
        
        return totals
    
    def extract_category_structure(self, json_data: Dict) -> List[Dict]:
        """
        Dynamically extract the category structure from QuickBooks JSON
        Returns a list of main categories with their subcategories and items
        """
        categories = []
        
        def get_section_name(row):
            """Extract section name from row header"""
            if 'Header' in row:
                header_col_data = row['Header'].get('ColData', [])
                if header_col_data:
                    return header_col_data[0].get('value', '').strip()
            return ""
        
        def get_data_items(rows_container):
            """Extract data item names from rows"""
            items = []
            if 'Row' in rows_container:
                for row in rows_container['Row']:
                    if row.get('type') == 'Data' or ('ColData' in row and 'Summary' not in row and 'Header' not in row):
                        col_data = row.get('ColData', [])
                        if col_data:
                            item_name = col_data[0].get('value', '').strip()
                            # Exclude only non-category items like fees and taxes
                            if item_name and item_name not in ['Interest & Credit Card Fees', 'SD - Excise Tax']:
                                items.append(item_name)
            return items
        
        # Process top-level rows (main categories)
        if 'Rows' in json_data and 'Row' in json_data['Rows']:
            for main_row in json_data['Rows']['Row']:
                if main_row.get('type') == 'Section':
                    main_name = get_section_name(main_row)
                    
                    # Skip non-category sections
                    if not main_name or main_name in ['Interest & Credit Card Fees', 'SD - Excise Tax']:
                        continue
                    
                    # Skip if it starts with "Total" (it's a summary row)
                    if main_name.startswith('Total'):
                        continue
                    
                    category = {
                        'key': main_name,
                        'name': main_name,
                        'subcategories': []
                    }
                    
                    # Process subcategories
                    if 'Rows' in main_row and 'Row' in main_row['Rows']:
                        for sub_row in main_row['Rows']['Row']:
                            if sub_row.get('type') == 'Section':
                                sub_name = get_section_name(sub_row)
                                
                                if not sub_name or sub_name.startswith('Total'):
                                    continue
                                
                                subcategory = {
                                    'key': sub_name,
                                    'name': f"   {sub_name}",
                                    'items': []
                                }
                                
                                # Extract items from subcategory
                                if 'Rows' in sub_row:
                                    subcategory['items'] = get_data_items(sub_row['Rows'])
                                
                                if subcategory['items']:  # Only add if has items
                                    category['subcategories'].append(subcategory)
                    
                    if category['subcategories']:  # Only add if has subcategories
                        categories.append(category)
        
        return categories
    
    def merge_category_structures(self, struct1: List[Dict], struct2: List[Dict]) -> List[Dict]:
        """
        Merge two category structures to capture all categories and items from both
        """
        # Create a dict for easy lookup
        merged = {}
        
        for cat in struct1 + struct2:
            cat_key = cat['key']
            if cat_key not in merged:
                merged[cat_key] = {
                    'key': cat['key'],
                    'name': cat['name'],
                    'subcategories': {}
                }
            
            for subcat in cat.get('subcategories', []):
                subcat_key = subcat['key']
                if subcat_key not in merged[cat_key]['subcategories']:
                    merged[cat_key]['subcategories'][subcat_key] = {
                        'key': subcat['key'],
                        'name': subcat['name'],
                        'items': set(subcat.get('items', []))
                    }
                else:
                    # Merge items
                    merged[cat_key]['subcategories'][subcat_key]['items'].update(subcat.get('items', []))
        
        # Convert back to list format
        result = []
        for cat_key in sorted(merged.keys()):
            cat = merged[cat_key]
            subcats = []
            for subcat_key in sorted(cat['subcategories'].keys()):
                subcat = cat['subcategories'][subcat_key]
                subcats.append({
                    'key': subcat['key'],
                    'name': subcat['name'],
                    'items': sorted(list(subcat['items']))
                })
            result.append({
                'key': cat['key'],
                'name': cat['name'],
                'subcategories': subcats
            })
        
        return result
    
    def generate_report(self, last_month_file: str, ytd_file: str, 
                       output_file: str, report_title: str, 
                       department_name: str, main_group_name: str = None) -> str:
        """
        Generate royalty volume calculation report matching original format EXACTLY
        
        Args:
            last_month_file: Path to last month JSON file
            ytd_file: Path to year-to-date JSON file
            output_file: Path for output Excel file
            report_title: Title for the report
            department_name: Department/location name
            main_group_name: Main group name for header (optional, defaults to department_name)
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Set main group name
            self.main_group_name = main_group_name if main_group_name else department_name
            
            # Load JSON data
            print(f"Loading {last_month_file}...")
            last_month_data = self.load_json_file(last_month_file)
            
            print(f"Loading {ytd_file}...")
            ytd_data = self.load_json_file(ytd_file)
            
            # Validate reports have data
            print("Validating report data...")
            if not self.validate_report_data(last_month_data):
                print("WARNING: Last month report indicates no data (NoReportData=true)")
            
            if not self.validate_report_data(ytd_data):
                print("WARNING: YTD report indicates no data (NoReportData=true)")
            
            # Get metadata
            lm_metadata = self.get_report_metadata(last_month_data)
            ytd_metadata = self.get_report_metadata(ytd_data)
            
            print(f"  Last Month: {lm_metadata['report_name']} ({lm_metadata['report_basis']} Basis)")
            print(f"  YTD: {ytd_metadata['report_name']} ({ytd_metadata['report_basis']} Basis)")
            
            # Extract totals
            print("Extracting data from QuickBooks JSON...")
            last_month_totals = self.extract_totals_from_json(last_month_data)
            ytd_totals = self.extract_totals_from_json(ytd_data)
            
            print(f"  Found {len(last_month_totals)} entries in last month data")
            print(f"  Found {len(ytd_totals)} entries in YTD data")
            
            # Get date information
            last_month_header = last_month_data.get('Header', {})
            ytd_header = ytd_data.get('Header', {})
            
            last_month_start = last_month_header.get('StartPeriod', '')
            last_month_end = last_month_header.get('EndPeriod', '')
            ytd_start = ytd_header.get('StartPeriod', '')
            ytd_end = ytd_header.get('EndPeriod', '')
            
            # Create workbook
            print("Creating Excel workbook...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales by Class Summary"
            
            # ===== ROW 1: Company Name =====
            ws.merge_cells('A1:E1')
            cell_a1 = ws['A1']
            cell_a1.value = self.main_group_name
            cell_a1.font = self.TITLE_FONT
            cell_a1.alignment = self.CENTER_ALIGN
            
            # ===== ROW 2: Report Title =====
            ws.merge_cells('A2:E2')
            cell_a2 = ws['A2']
            cell_a2.value = f"{report_title}"
            cell_a2.font = self.SUBTITLE_FONT
            cell_a2.alignment = self.CENTER_ALIGN
            
            # ===== ROW 3: Date =====
            ws.merge_cells('A3:E3')
            cell_a3 = ws['A3']
            # Format date from last_month_end (e.g., "2025-10-31" -> "October 2025")
            try:
                date_obj = datetime.strptime(last_month_end, '%Y-%m-%d')
                date_str = date_obj.strftime('%B %Y')
            except Exception as e:
                date_str = last_month_end
            cell_a3.value = date_str
            cell_a3.font = self.DATE_FONT
            cell_a3.alignment = self.CENTER_ALIGN
            
            # ===== ROW 4: Empty =====
            
            # ===== ROW 5: Location Headers (merged) =====
            ws.merge_cells('B5:C5')
            cell_b5 = ws['B5']  # Use top-left cell of merged range
            cell_b5.value = department_name
            cell_b5.font = self.HEADER_FONT
            cell_b5.alignment = self.CENTER_ALIGN
            
            ws.merge_cells('D5:E5')  # Use top-left cell of merged range
            cell_d5 = ws['D5']  # Use top-left cell of merged range
            cell_d5.value = "Total"
            cell_d5.font = self.HEADER_FONT
            cell_d5.alignment = self.CENTER_ALIGN
            
            # ===== ROW 6: Period Headers =====
            cell_b6 = ws['B6']
            try:
                lm_start_obj = datetime.strptime(last_month_start, '%Y-%m-%d')
                lm_end_obj = datetime.strptime(last_month_end, '%Y-%m-%d')
                # Format: "Oct 1 - Oct 31, 2025" (with days and year)
                b6_str = f"{lm_start_obj.strftime('%b')} {lm_start_obj.day} - {lm_end_obj.strftime('%b')} {lm_end_obj.day}, {lm_end_obj.year}"
            except:
                b6_str = last_month_start
            cell_b6.value = b6_str
            cell_b6.font = self.HEADER_FONT
            cell_b6.alignment = self.CENTER_ALIGN
            
            cell_c6 = ws['C6']
            try:
                ytd_start_obj = datetime.strptime(ytd_start, '%Y-%m-%d')
                ytd_end_obj = datetime.strptime(ytd_end, '%Y-%m-%d')
                # Format: "Jan 1 - Oct 31, 2025 (YTD)" (with days)
                c6_str = f"{ytd_start_obj.strftime('%b')} {ytd_start_obj.day} - {ytd_end_obj.strftime('%b')} {ytd_end_obj.day}, {ytd_end_obj.year} (YTD)"
            except:
                c6_str = f"{ytd_start} - {ytd_end} (YTD)"
            cell_c6.value = c6_str
            cell_c6.font = self.HEADER_FONT
            cell_c6.alignment = self.CENTER_ALIGN
            
            cell_d6 = ws['D6']
            cell_d6.value = b6_str
            cell_d6.font = self.HEADER_FONT
            cell_d6.alignment = self.CENTER_ALIGN
            
            cell_e6 = ws['E6']
            cell_e6.value = c6_str
            cell_e6.font = self.HEADER_FONT
            cell_e6.alignment = self.CENTER_ALIGN
            
            # ===== DATA ROWS =====
            current_row = 7
            
            # Dynamically extract category structure from JSON
            # Merge structures from both files to ensure all categories/items are captured
            print("Extracting category structure from JSON data...")
            lm_categories = self.extract_category_structure(last_month_data)
            ytd_categories = self.extract_category_structure(ytd_data)
            
            # Use the structure with more data (usually YTD has more complete structure)
            categories = ytd_categories if len(ytd_categories) >= len(lm_categories) else lm_categories
            
            # Merge any missing items from the other structure
            categories = self.merge_category_structures(lm_categories, ytd_categories)
            
            print(f"  Found {len(categories)} main categories")
            
            # Write data rows with exact format
            for category in categories:
                # Main category header
                ws.cell(row=current_row, column=1).value = category['name']
                ws.cell(row=current_row, column=1).font = self.CATEGORY_FONT
                ws.cell(row=current_row, column=1).alignment = self.LEFT_ALIGN
                
                # Formula cells for main category
                ws.cell(row=current_row, column=4).value = f"=B{current_row}"
                ws.cell(row=current_row, column=4).number_format = '#,##0.00\\ _Ç'
                ws.cell(row=current_row, column=4).font = self.DATA_FONT
                ws.cell(row=current_row, column=4).alignment = self.RIGHT_ALIGN
                
                ws.cell(row=current_row, column=5).value = f"=C{current_row}"
                ws.cell(row=current_row, column=5).number_format = '#,##0.00\\ _Ç'
                ws.cell(row=current_row, column=5).font = self.DATA_FONT
                ws.cell(row=current_row, column=5).alignment = self.RIGHT_ALIGN
                
                current_row += 1
                
                # Subcategories
                for subcat in category['subcategories']:
                    # Subcategory header
                    ws.cell(row=current_row, column=1).value = subcat['name']
                    ws.cell(row=current_row, column=1).font = self.CATEGORY_FONT
                    ws.cell(row=current_row, column=1).alignment = self.LEFT_ALIGN
                    
                    # Check for Header-level value (unclassified amounts at subcategory level)
                    subcat_path = f"{category['key']}::{subcat['key']}"
                    header_key = f"HEADER::{subcat_path}"
                    lm_header_val = last_month_totals.get(header_key, 0.0)
                    ytd_header_val = ytd_totals.get(header_key, 0.0)
                    
                    # Write Header-level values to subcategory row if they exist
                    if lm_header_val != 0:
                        ws.cell(row=current_row, column=2).value = lm_header_val
                        ws.cell(row=current_row, column=2).number_format = '#,##0.00\\ _Ç'
                        ws.cell(row=current_row, column=2).font = self.DATA_FONT
                        ws.cell(row=current_row, column=2).alignment = self.RIGHT_ALIGN
                    
                    if ytd_header_val != 0:
                        ws.cell(row=current_row, column=3).value = ytd_header_val
                        ws.cell(row=current_row, column=3).number_format = '#,##0.00\\ _Ç'
                        ws.cell(row=current_row, column=3).font = self.DATA_FONT
                        ws.cell(row=current_row, column=3).alignment = self.RIGHT_ALIGN
                    
                    ws.cell(row=current_row, column=4).value = f"=B{current_row}"
                    ws.cell(row=current_row, column=4).number_format = '#,##0.00\\ _Ç'
                    ws.cell(row=current_row, column=4).font = self.DATA_FONT
                    ws.cell(row=current_row, column=4).alignment = self.RIGHT_ALIGN
                    
                    ws.cell(row=current_row, column=5).value = f"=C{current_row}"
                    ws.cell(row=current_row, column=5).number_format = '#,##0.00\\ _Ç'
                    ws.cell(row=current_row, column=5).font = self.DATA_FONT
                    ws.cell(row=current_row, column=5).alignment = self.RIGHT_ALIGN
                    
                    current_row += 1
                    
                    # Items (actual data rows)
                    for item in subcat['items']:
                        ws.cell(row=current_row, column=1).value = f"      {item}"
                        ws.cell(row=current_row, column=1).font = self.CATEGORY_FONT
                        ws.cell(row=current_row, column=1).alignment = self.LEFT_ALIGN
                        
                        # Look up actual values using FULL PATH: MainCategory::SubCategory::Item
                        full_path = f"{category['key']}::{subcat['key']}::{item}"
                        data_key = f"DATA::{full_path}"
                        lm_value = last_month_totals.get(data_key, 0.0)
                        ytd_value = ytd_totals.get(data_key, 0.0)
                        
                        # Write last month value if exists
                        if lm_value and lm_value != 0:
                            ws.cell(row=current_row, column=2).value = lm_value
                            ws.cell(row=current_row, column=2).number_format = '#,##0.00\\ _Ç'
                            ws.cell(row=current_row, column=2).font = self.DATA_FONT
                            ws.cell(row=current_row, column=2).alignment = self.RIGHT_ALIGN
                        
                        # Write YTD value if exists
                        if ytd_value and ytd_value != 0:
                            ws.cell(row=current_row, column=3).value = ytd_value
                            ws.cell(row=current_row, column=3).number_format = '#,##0.00\\ _Ç'
                            ws.cell(row=current_row, column=3).font = self.DATA_FONT
                            ws.cell(row=current_row, column=3).alignment = self.RIGHT_ALIGN
                        
                        ws.cell(row=current_row, column=4).value = f"=B{current_row}"
                        ws.cell(row=current_row, column=4).number_format = '#,##0.00\\ _Ç'
                        ws.cell(row=current_row, column=4).font = self.DATA_FONT
                        ws.cell(row=current_row, column=4).alignment = self.RIGHT_ALIGN
                        
                        ws.cell(row=current_row, column=5).value = f"=C{current_row}"
                        ws.cell(row=current_row, column=5).number_format = '#,##0.00\\ _Ç'
                        ws.cell(row=current_row, column=5).font = self.DATA_FONT
                        ws.cell(row=current_row, column=5).alignment = self.RIGHT_ALIGN
                        
                        current_row += 1
                    
                    # Subcategory total
                    total_key = f"Total {subcat['key']}"
                    total_name = f"   Total {subcat['key']}"
                    
                    ws.cell(row=current_row, column=1).value = total_name
                    ws.cell(row=current_row, column=1).font = self.TOTAL_FONT
                    ws.cell(row=current_row, column=1).alignment = self.LEFT_ALIGN
                    
                    # Get actual total from JSON
                    lm_total = last_month_totals.get(total_key, 0.0)
                    ytd_total = ytd_totals.get(total_key, 0.0)
                    
                    ws.cell(row=current_row, column=2).value = lm_total
                    ws.cell(row=current_row, column=2).number_format = '"$"* #,##0.00\\ _Ç'
                    ws.cell(row=current_row, column=2).font = self.TOTAL_FONT
                    ws.cell(row=current_row, column=2).alignment = self.RIGHT_ALIGN
                    
                    ws.cell(row=current_row, column=3).value = ytd_total
                    ws.cell(row=current_row, column=3).number_format = '"$"* #,##0.00\\ _Ç'
                    ws.cell(row=current_row, column=3).font = self.TOTAL_FONT
                    ws.cell(row=current_row, column=3).alignment = self.RIGHT_ALIGN
                    
                    ws.cell(row=current_row, column=4).value = f"=B{current_row}"
                    ws.cell(row=current_row, column=4).number_format = '"$"* #,##0.00\\ _Ç'
                    ws.cell(row=current_row, column=4).font = self.TOTAL_FONT
                    ws.cell(row=current_row, column=4).alignment = self.RIGHT_ALIGN
                    
                    ws.cell(row=current_row, column=5).value = f"=C{current_row}"
                    ws.cell(row=current_row, column=5).number_format = '"$"* #,##0.00\\ _Ç'
                    ws.cell(row=current_row, column=5).font = self.TOTAL_FONT
                    ws.cell(row=current_row, column=5).alignment = self.RIGHT_ALIGN
                    
                    current_row += 1
                
                # Main category total
                main_total_key = f"Total {category['key']}"
                main_total_name = f"Total {category['key']}"
                
                ws.cell(row=current_row, column=1).value = main_total_name
                ws.cell(row=current_row, column=1).font = self.TOTAL_FONT
                ws.cell(row=current_row, column=1).alignment = self.LEFT_ALIGN
                
                # Get actual total from JSON
                lm_main_total = last_month_totals.get(main_total_key, 0.0)
                ytd_main_total = ytd_totals.get(main_total_key, 0.0)
                
                ws.cell(row=current_row, column=2).value = lm_main_total
                ws.cell(row=current_row, column=2).number_format = '"$"* #,##0.00\\ _Ç'
                ws.cell(row=current_row, column=2).font = self.TOTAL_FONT
                ws.cell(row=current_row, column=2).alignment = self.RIGHT_ALIGN
                
                ws.cell(row=current_row, column=3).value = ytd_main_total
                ws.cell(row=current_row, column=3).number_format = '"$"* #,##0.00\\ _Ç'
                ws.cell(row=current_row, column=3).font = self.TOTAL_FONT
                ws.cell(row=current_row, column=3).alignment = self.RIGHT_ALIGN
                
                ws.cell(row=current_row, column=4).value = f"=B{current_row}"
                ws.cell(row=current_row, column=4).number_format = '"$"* #,##0.00\\ _Ç'
                ws.cell(row=current_row, column=4).font = self.TOTAL_FONT
                ws.cell(row=current_row, column=4).alignment = self.RIGHT_ALIGN
                
                ws.cell(row=current_row, column=5).value = f"=C{current_row}"
                ws.cell(row=current_row, column=5).number_format = '"$"* #,##0.00\\ _Ç'
                ws.cell(row=current_row, column=5).font = self.TOTAL_FONT
                ws.cell(row=current_row, column=5).alignment = self.RIGHT_ALIGN
                
                current_row += 1
            
            # Grand Total
            ws.cell(row=current_row, column=1).value = "TOTAL"
            ws.cell(row=current_row, column=1).font = self.TOTAL_FONT
            ws.cell(row=current_row, column=1).alignment = self.LEFT_ALIGN
            
            # Get grand totals from JSON
            lm_grand_total = last_month_totals.get('TOTAL', 0.0)
            ytd_grand_total = ytd_totals.get('TOTAL', 0.0)
            
            ws.cell(row=current_row, column=2).value = lm_grand_total
            ws.cell(row=current_row, column=2).number_format = '"$"* #,##0.00\\ _Ç'
            ws.cell(row=current_row, column=2).font = self.TOTAL_FONT
            ws.cell(row=current_row, column=2).alignment = self.RIGHT_ALIGN
            
            ws.cell(row=current_row, column=3).value = ytd_grand_total
            ws.cell(row=current_row, column=3).number_format = '"$"* #,##0.00\\ _Ç'
            ws.cell(row=current_row, column=3).font = self.TOTAL_FONT
            ws.cell(row=current_row, column=3).alignment = self.RIGHT_ALIGN
            
            ws.cell(row=current_row, column=4).value = f"=B{current_row}"
            ws.cell(row=current_row, column=4).number_format = '"$"* #,##0.00\\ _Ç'
            ws.cell(row=current_row, column=4).font = self.TOTAL_FONT
            ws.cell(row=current_row, column=4).alignment = self.RIGHT_ALIGN
            
            ws.cell(row=current_row, column=5).value = f"=C{current_row}"
            ws.cell(row=current_row, column=5).number_format = '"$"* #,##0.00\\ _Ç'
            ws.cell(row=current_row, column=5).font = self.TOTAL_FONT
            ws.cell(row=current_row, column=5).alignment = self.RIGHT_ALIGN
            
            current_row += 1
            
            # ===== FOOTER =====
            # Add empty row
            current_row += 2
            
            # Footer with real-time timestamp in EST and basis
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell_footer = ws[f'A{current_row}']
            
            # Get basis from last month data
            basis = last_month_header.get('ReportBasis', 'Cash')
            
            # Generate real-time timestamp in EST timezone
            try:
                from zoneinfo import ZoneInfo
                est_tz = ZoneInfo('America/New_York')
                now_est = datetime.now(est_tz)
            except ImportError:
                # Fallback for Python < 3.9: use pytz or manual offset
                try:
                    import pytz
                    est_tz = pytz.timezone('America/New_York')
                    now_est = datetime.now(est_tz)
                except ImportError:
                    # Manual EST offset (UTC-5) as last resort
                    from datetime import timezone, timedelta
                    est_offset = timezone(timedelta(hours=-5))
                    now_est = datetime.now(est_offset)
            
            # Format: "Friday, Dec 5, 2025 10:30:45 AM EST - Cash Basis"
            day_name = now_est.strftime('%A')
            month_name = now_est.strftime('%b')
            day = now_est.day
            year = now_est.year
            time_12hr = now_est.strftime('%I:%M:%S %p').lstrip('0')
            
            footer_text = f"{basis} Basis {day_name}, {month_name} {day}, {year} {time_12hr} EST GMT-05:00"
            
            cell_footer.value = footer_text
            cell_footer.font = Font(name='Arial', size=8, bold=False)
            cell_footer.alignment = self.CENTER_ALIGN
            
            # Set column widths to match original exactly
            ws.column_dimensions['A'].width = 33.5
            ws.column_dimensions['B'].width = 31.8
            ws.column_dimensions['C'].width = 31.8
            ws.column_dimensions['D'].width = 31.8
            ws.column_dimensions['E'].width = 31.8
            
            # Save workbook
            print(f"Saving report to {output_file}...")
            wb.save(output_file)
            
            print(f"[SUCCESS] Report generated successfully: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"[ERROR] Error generating report: {str(e)}")
            raise
    
    def convert_to_pdf(self, excel_file: str, pdf_file: str = None) -> str:
        """
        Convert Excel file to PDF. Tries LibreOffice first, then falls back to win32com on Windows.
        
        Args:
            excel_file: Path to the Excel file
            pdf_file: Path for output PDF file (optional, defaults to same name with .pdf extension)
            
        Returns:
            Path to generated PDF file
            
        Note:
            Requires one of:
            - LibreOffice: sudo apt-get install libreoffice (Linux) or download from libreoffice.org
            - Microsoft Excel (Windows only)
        """
        # Determine PDF output path
        if pdf_file is None:
            pdf_file = os.path.splitext(excel_file)[0] + '.pdf'
        
        print(f"Converting {excel_file} to PDF...")
        
        errors = []
        
        # Try LibreOffice first (cross-platform)
        try:
            return self._convert_to_pdf_libreoffice(excel_file, pdf_file)
        except Exception as e:
            errors.append(f"LibreOffice: {str(e)}")
            print(f"[INFO] LibreOffice conversion failed, trying alternative...")
        
        # Try win32com on Windows as fallback
        try:
            return self._convert_to_pdf_win32(excel_file, pdf_file)
        except Exception as e:
            errors.append(f"Win32com/Excel: {str(e)}")
        
        # Both methods failed
        error_message = (
            "PDF conversion failed. Tried methods:\n" +
            "\n".join(f"  - {err}" for err in errors) +
            "\n\nPlease install one of:\n"
            "  - LibreOffice: sudo apt-get install libreoffice (Linux) or https://www.libreoffice.org/\n"
            "  - Microsoft Excel (Windows only)"
        )
        print(f"[ERROR] {error_message}")
        raise RuntimeError(error_message)
    
    def _convert_to_pdf_libreoffice(self, excel_file: str, pdf_file: str) -> str:
        """
        Convert Excel to PDF using LibreOffice (cross-platform)
        """
        import subprocess
        import shutil
        
        # Get absolute paths
        excel_path = os.path.abspath(excel_file)
        output_dir = os.path.dirname(os.path.abspath(pdf_file))
        
        # Find LibreOffice executable
        libreoffice_paths = [
            'libreoffice',  # Linux (in PATH)
            'soffice',  # Alternative Linux
            '/usr/bin/libreoffice',  # Linux default
            '/usr/bin/soffice',  # Linux alternative
            '/Applications/LibreOffice.app/Contents/MacOS/soffice',  # macOS
            'C:\\Program Files\\LibreOffice\\program\\soffice.exe',  # Windows
            'C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe',  # Windows 32-bit
        ]
        
        libreoffice_cmd = None
        for path in libreoffice_paths:
            if shutil.which(path) or os.path.exists(path):
                libreoffice_cmd = path
                break
        
        if libreoffice_cmd is None:
            raise RuntimeError("LibreOffice not found")
        
        # Run LibreOffice in headless mode to convert to PDF
        cmd = [
            libreoffice_cmd,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            excel_path
        ]
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=120  # 2 minute timeout
        )
        
        if result.returncode != 0:
            error_msg = result.stderr or result.stdout or "Unknown error"
            raise RuntimeError(f"LibreOffice conversion failed: {error_msg}")
        
        # LibreOffice outputs PDF with same base name in output directory
        expected_pdf = os.path.join(
            output_dir, 
            os.path.splitext(os.path.basename(excel_file))[0] + '.pdf'
        )
        
        # Rename if needed to match requested pdf_file path
        if expected_pdf != os.path.abspath(pdf_file):
            if os.path.exists(expected_pdf):
                shutil.move(expected_pdf, pdf_file)
        
        if not os.path.exists(pdf_file):
            raise RuntimeError(f"PDF file was not created: {pdf_file}")
        
        print(f"[SUCCESS] PDF generated successfully (LibreOffice): {pdf_file}")
        return pdf_file
    
    def _convert_to_pdf_win32(self, excel_file: str, pdf_file: str) -> str:
        """
        Convert Excel to PDF using win32com (Windows only, requires Microsoft Excel)
        """
        import time
        
        try:
            import win32com.client
        except ImportError:
            raise RuntimeError("win32com not available (not Windows or pywin32 not installed)")
        
        # Get absolute paths
        excel_path = os.path.abspath(excel_file)
        pdf_path = os.path.abspath(pdf_file)
        
        # Create Excel application instance
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            # Open workbook
            workbook = excel.Workbooks.Open(excel_path)
            
            # Get the active sheet
            sheet = workbook.ActiveSheet
            
            # Set page setup for better PDF output
            sheet.PageSetup.Orientation = 2  # Landscape
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False
            sheet.PageSetup.Zoom = False
            
            # Export to PDF (0 = xlTypePDF)
            workbook.ExportAsFixedFormat(0, pdf_path)
            
            print(f"[SUCCESS] PDF generated successfully (Excel/win32): {pdf_file}")
            
        finally:
            # Close workbook and quit Excel
            try:
                workbook.Close(SaveChanges=False)
            except:
                pass
            excel.Quit()
            
            # Release COM objects
            del workbook
            del excel
            
            # Small delay to ensure Excel process terminates
            time.sleep(0.5)
        
        return pdf_file
    
    def generate_report_with_pdf(self, last_month_file: str, ytd_file: str, 
                                  output_file: str, report_title: str, 
                                  department_name: str, main_group_name: str = None) -> Tuple[str, str]:
        """
        Generate royalty report and convert to PDF
        
        Args:
            last_month_file: Path to last month JSON file
            ytd_file: Path to year-to-date JSON file
            output_file: Path for output Excel file
            report_title: Title for the report
            department_name: Department/location name
            main_group_name: Main group name for header (optional)
            
        Returns:
            Tuple of (excel_path, pdf_path)
        """
        # Generate Excel report
        excel_path = self.generate_report(
            last_month_file=last_month_file,
            ytd_file=ytd_file,
            output_file=output_file,
            report_title=report_title,
            department_name=department_name,
            main_group_name=main_group_name
        )
        
        # Convert to PDF
        pdf_path = self.convert_to_pdf(excel_path)
        
        return excel_path, pdf_path


def main():
    """Main function to demonstrate usage"""
    
    # Example usage
    generator = RoyaltyReportGenerator()
    
    # Define file paths
    last_month_file = "Department/SERVPRO Team Marchese/id7rvcr.json"  # Result from Last Month api call
    ytd_file = "Department/SERVPRO Team Marchese/id7rvcrytd.json"  # Result from YTD api call
    output_file = "Department/Generated_RVCR_Sarpy County_11533.xlsx"
    
    # Report configuration
    report_title = "RVCR - Sarpy County 11533" # QBO Department Name
    department_name = "Sarpy County 11533" # QBO Department Name
    main_group_name = "SERVPRO of Sooland, Lincoln, Sarpy County, Kearney & North Platte, Grand Island & Hastings" # QBO Company Name
    
    # Generate report with PDF
    try:
        excel_path, pdf_path = generator.generate_report_with_pdf(
            last_month_file=last_month_file,
            ytd_file=ytd_file,
            output_file=output_file,
            report_title=report_title,
            department_name=department_name,
            main_group_name=main_group_name
        )
        
        print("\n" + "="*60)
        print("Report Generation Complete!")
        print("="*60)
        print(f"Excel file: {excel_path}")
        print(f"PDF file: {pdf_path}")
        
    except Exception as e:
        print(f"\n[ERROR] Failed to generate report: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
