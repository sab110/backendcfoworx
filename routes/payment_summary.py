"""
Payment Summary (Royalty Report) Generation Endpoint
Generates royalty calculation reports based on RVCR data
"""

import os
import json
import uuid
import tempfile
from datetime import datetime, timedelta
from typing import Optional
from dateutil.relativedelta import relativedelta

from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from models import (
    QuickBooksToken, 
    CompanyInfo, 
    CompanyLicenseMapping, 
    GeneratedReport
)
from services.azure_storage_service import AzureStorageService
from routes.quickbooks_auth import refresh_qbo_token
from routes.rvcr_reports import (
    fetch_class_sales_report,
    query_class_ids,
    get_qbo_base_url
)

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()


# ============================================================================
# ROYALTY CALCULATION CONSTANTS
# ============================================================================

# Standard Rate Royalty Tiers (excluding Subcontract and Reconstruction)
STANDARD_RATE_TIERS = [
    {"max": 12849.99, "rate": 0.10, "fixed": 0, "fixed_fee": 45},
    {"max": 21416.99, "rate": 0.09, "fixed": 0, "fixed_fee": 65},
    {"max": 32124.99, "rate": 0.08, "fixed": 0, "fixed_fee": 85},
    {"max": 42832.99, "rate": 0.075, "fixed": 0, "fixed_fee": 95},
    {"max": 74956.99, "rate": 0.07, "fixed": 0, "fixed_fee": 115},
    {"max": 107082.99, "rate": 0.065, "fixed": 5247.00, "threshold": 74956.99, "fixed_fee": 115},
    {"max": 160623.99, "rate": 0.06, "fixed": 7335.00, "threshold": 107082.99, "fixed_fee": 115},
    {"max": 214164.99, "rate": 0.055, "fixed": 10547.00, "threshold": 160623.99, "fixed_fee": 115},
    {"max": float('inf'), "rate": 0.05, "fixed": 13492.00, "threshold": 214164.99, "fixed_fee": 115},
]

# Reduced Rate Royalty (Subcontract and Reconstruction)
REDUCED_RATE_THRESHOLD = 32124.99
REDUCED_RATE_BELOW = 0.04  # 4% for amounts up to threshold
REDUCED_RATE_ABOVE = 0.03  # 3% for amounts above threshold
REDUCED_RATE_FIXED = 1285.00  # Fixed amount when above threshold

# Fee Rates
NATIONAL_ACCOUNTS_FEE_RATE = 0.005  # 0.5%
NATIONAL_BRAND_FUND_RATE = 0.025  # 2.5%
NATIONAL_BRAND_FUND_CAP_2025 = 1450000  # $1,450,000 cap for 2025
NATIONAL_BRAND_FUND_REDUCED_RATE = 0.0025  # 0.25%


# ============================================================================
# ROYALTY CALCULATION FUNCTIONS
# ============================================================================

def calculate_standard_rate_royalty(volume: float) -> dict:
    """
    Calculate royalty for Standard Rate Services using tiered table.
    Returns dict with royalty amount, tier info, and description.
    """
    if volume <= 0:
        return {
            "royalty": 0,
            "fixed_fee": 45,
            "description": "No revenue",
            "rate": 0,
            "tier_description": "$0"
        }
    
    for tier in STANDARD_RATE_TIERS:
        if volume <= tier["max"]:
            if "threshold" in tier:
                # Tiered calculation: fixed + rate on excess
                excess = volume - tier["threshold"]
                royalty = tier["fixed"] + (excess * tier["rate"])
                description = f"${tier['fixed']:,.2f} + {tier['rate']*100:.2f}% of ${excess:,.2f}"
                tier_desc = f"${tier['threshold']:,.2f}+"
            else:
                # Simple percentage
                royalty = volume * tier["rate"]
                description = f"{tier['rate']*100:.1f}% of ${volume:,.2f}"
                tier_desc = f"$0 - ${tier['max']:,.2f}"
            
            return {
                "royalty": round(royalty, 2),
                "fixed_fee": tier["fixed_fee"],
                "description": description,
                "rate": tier["rate"],
                "tier_description": tier_desc
            }
    
    # Should not reach here, but just in case
    return {"royalty": 0, "fixed_fee": 115, "description": "Error", "rate": 0, "tier_description": "N/A"}


def calculate_reduced_rate_royalty(volume: float) -> dict:
    """
    Calculate royalty for Reduced Rate Services (Subcontract + Reconstruction).
    4% up to $32,124.99, then 3% on excess.
    """
    if volume <= 0:
        return {
            "royalty": 0,
            "description": "No revenue",
            "excess_amount": 0
        }
    
    if volume <= REDUCED_RATE_THRESHOLD:
        royalty = volume * REDUCED_RATE_BELOW
        return {
            "royalty": round(royalty, 2),
            "description": f"{REDUCED_RATE_BELOW*100:.0f}% of ${volume:,.2f}",
            "excess_amount": 0
        }
    else:
        excess = volume - REDUCED_RATE_THRESHOLD
        royalty = REDUCED_RATE_FIXED + (excess * REDUCED_RATE_ABOVE)
        return {
            "royalty": round(royalty, 2),
            "description": f"${REDUCED_RATE_FIXED:,.2f} + {REDUCED_RATE_ABOVE*100:.0f}% of ${excess:,.2f}",
            "excess_amount": round(excess, 2)
        }


def calculate_national_brand_fund_fee(
    standard_rate_this_month: float,
    standard_rate_ytd_before: float,
    cap: float = NATIONAL_BRAND_FUND_CAP_2025
) -> dict:
    """
    Calculate National Brand Fund Fee with annual cap.
    2.5% of Standard Rate Services, capped at $1,450,000/year (2025).
    """
    if standard_rate_this_month <= 0:
        return {
            "fee": 0,
            "description": "No standard rate revenue",
            "capped": False,
            "remaining_cap": max(0, cap - standard_rate_ytd_before)
        }
    
    remaining_cap = max(0, cap - standard_rate_ytd_before)
    
    if remaining_cap <= 0:
        # Cap already reached
        return {
            "fee": 0,
            "description": "Annual cap reached",
            "capped": True,
            "remaining_cap": 0
        }
    
    # Calculate fee on the lesser of this month's amount or remaining cap
    applicable_amount = min(standard_rate_this_month, remaining_cap)
    fee = applicable_amount * NATIONAL_BRAND_FUND_RATE
    
    capped = applicable_amount < standard_rate_this_month
    
    return {
        "fee": round(fee, 2),
        "description": f"{NATIONAL_BRAND_FUND_RATE*100:.1f}% of ${applicable_amount:,.2f}",
        "capped": capped,
        "remaining_cap": round(remaining_cap - applicable_amount, 2) if not capped else 0
    }


# ============================================================================
# DATA EXTRACTION FROM RVCR JSON
# ============================================================================

def extract_category_totals(rvcr_data: dict) -> dict:
    """
    Extract category totals from RVCR JSON data.
    Returns dict with all category values for Commercial/Residential.
    """
    results = {
        "water": {"commercial": 0, "residential": 0},
        "fire": {"commercial": 0, "residential": 0},
        "mold_bio": {"commercial": 0, "residential": 0},
        "other": {"commercial": 0, "residential": 0},
        "subcontract": {"commercial": 0, "residential": 0},
        "reconstruction": {"commercial": 0, "residential": 0},
    }
    
    # Category mapping from RVCR column titles
    category_map = {
        "Total Commercial - Water": ("water", "commercial"),
        "Total Residential - Water": ("water", "residential"),
        "Total Commercial - Fire": ("fire", "commercial"),
        "Total Residential - Fire": ("fire", "residential"),
        "Total Commercial - Mold/Bio Hazard": ("mold_bio", "commercial"),
        "Total Residential - Mold/Bio Hazard": ("mold_bio", "residential"),
        "Total Commercial - Other": ("other", "commercial"),
        "Total Residential - Other": ("other", "residential"),
        "Total Commercial - Subcontract": ("subcontract", "commercial"),
        "Total Residential - Subcontract": ("subcontract", "residential"),
        "Total Commercial - Reconstruction": ("reconstruction", "commercial"),
        "Total Residential - Reconstruction": ("reconstruction", "residential"),
    }
    
    # Get column definitions
    columns = rvcr_data.get("Columns", {}).get("Column", [])
    col_titles = [col.get("ColTitle", "") for col in columns]
    
    # Find TOTAL row
    rows = rvcr_data.get("Rows", {}).get("Row", [])
    total_row = None
    for row in rows:
        if row.get("type") == "Section" and row.get("group") == "GrandTotal":
            summary = row.get("Summary", {})
            total_row = summary.get("ColData", [])
            break
    
    if not total_row:
        print("⚠️ Could not find TOTAL row in RVCR data")
        return results
    
    # Extract values for each category
    for i, col_title in enumerate(col_titles):
        if col_title in category_map:
            category, type_key = category_map[col_title]
            if i < len(total_row):
                value_str = total_row[i].get("value", "0")
                try:
                    value = float(value_str) if value_str else 0
                    results[category][type_key] = value
                except ValueError:
                    results[category][type_key] = 0
    
    return results


def calculate_payment_summary(
    last_month_data: dict,
    ytd_data: dict,
    franchise_number: str,
    department_name: str,
    owner_name: str = "",
    period_month: int = None,
    period_year: int = None
) -> dict:
    """
    Calculate complete payment summary from RVCR data.
    """
    # Extract category totals for this month
    categories = extract_category_totals(last_month_data)
    
    # Extract YTD totals
    ytd_categories = extract_category_totals(ytd_data)
    
    # Calculate subtotals - This Month
    water_total = categories["water"]["commercial"] + categories["water"]["residential"]
    fire_total = categories["fire"]["commercial"] + categories["fire"]["residential"]
    mold_total = categories["mold_bio"]["commercial"] + categories["mold_bio"]["residential"]
    other_total = categories["other"]["commercial"] + categories["other"]["residential"]
    subcontract_total = categories["subcontract"]["commercial"] + categories["subcontract"]["residential"]
    reconstruction_total = categories["reconstruction"]["commercial"] + categories["reconstruction"]["residential"]
    
    standard_rate_total = water_total + fire_total + mold_total + other_total
    reduced_rate_total = subcontract_total + reconstruction_total
    grand_total = standard_rate_total + reduced_rate_total
    
    # Calculate subtotals - YTD
    ytd_water = ytd_categories["water"]["commercial"] + ytd_categories["water"]["residential"]
    ytd_fire = ytd_categories["fire"]["commercial"] + ytd_categories["fire"]["residential"]
    ytd_mold = ytd_categories["mold_bio"]["commercial"] + ytd_categories["mold_bio"]["residential"]
    ytd_other = ytd_categories["other"]["commercial"] + ytd_categories["other"]["residential"]
    ytd_subcontract = ytd_categories["subcontract"]["commercial"] + ytd_categories["subcontract"]["residential"]
    ytd_reconstruction = ytd_categories["reconstruction"]["commercial"] + ytd_categories["reconstruction"]["residential"]
    
    ytd_standard_rate = ytd_water + ytd_fire + ytd_mold + ytd_other
    ytd_reduced_rate = ytd_subcontract + ytd_reconstruction
    ytd_grand_total = ytd_standard_rate + ytd_reduced_rate
    
    # Calculate YTD before this month (for cap calculations)
    ytd_standard_before = ytd_standard_rate - standard_rate_total
    
    # Calculate royalties
    standard_royalty = calculate_standard_rate_royalty(standard_rate_total)
    reduced_royalty = calculate_reduced_rate_royalty(reduced_rate_total)
    
    # Calculate fees
    fixed_fee = standard_royalty["fixed_fee"]
    national_accounts_fee = round(standard_rate_total * NATIONAL_ACCOUNTS_FEE_RATE, 2)
    
    brand_fund = calculate_national_brand_fund_fee(
        standard_rate_total,
        ytd_standard_before
    )
    national_brand_fee = brand_fund["fee"]
    
    national_brand_reduced_fee = round(reduced_rate_total * NATIONAL_BRAND_FUND_REDUCED_RATE, 2)
    
    # Total payable
    total_royalty_payable = (
        standard_royalty["royalty"] + 
        reduced_royalty["royalty"]
    )
    
    total_fees_payable = (
        fixed_fee +
        national_accounts_fee +
        national_brand_fee +
        national_brand_reduced_fee
    )
    
    grand_total_payable = total_royalty_payable + total_fees_payable
    
    return {
        "franchise_number": franchise_number,
        "department_name": department_name,
        "owner_name": owner_name,
        "period_month": period_month,
        "period_year": period_year,
        "categories": {
            "water": {
                "commercial": {"this_month": categories["water"]["commercial"], "ytd": ytd_categories["water"]["commercial"]},
                "residential": {"this_month": categories["water"]["residential"], "ytd": ytd_categories["water"]["residential"]},
            },
            "fire": {
                "commercial": {"this_month": categories["fire"]["commercial"], "ytd": ytd_categories["fire"]["commercial"]},
                "residential": {"this_month": categories["fire"]["residential"], "ytd": ytd_categories["fire"]["residential"]},
            },
            "mold_bio": {
                "commercial": {"this_month": categories["mold_bio"]["commercial"], "ytd": ytd_categories["mold_bio"]["commercial"]},
                "residential": {"this_month": categories["mold_bio"]["residential"], "ytd": ytd_categories["mold_bio"]["residential"]},
            },
            "other": {
                "commercial": {"this_month": categories["other"]["commercial"], "ytd": ytd_categories["other"]["commercial"]},
                "residential": {"this_month": categories["other"]["residential"], "ytd": ytd_categories["other"]["residential"]},
            },
            "subcontract": {
                "commercial": {"this_month": categories["subcontract"]["commercial"], "ytd": ytd_categories["subcontract"]["commercial"]},
                "residential": {"this_month": categories["subcontract"]["residential"], "ytd": ytd_categories["subcontract"]["residential"]},
            },
            "reconstruction": {
                "commercial": {"this_month": categories["reconstruction"]["commercial"], "ytd": ytd_categories["reconstruction"]["commercial"]},
                "residential": {"this_month": categories["reconstruction"]["residential"], "ytd": ytd_categories["reconstruction"]["residential"]},
            },
        },
        "subtotals": {
            "standard_rate": {"this_month": standard_rate_total, "ytd": ytd_standard_rate},
            "reduced_rate": {"this_month": reduced_rate_total, "ytd": ytd_reduced_rate},
            "total": {"this_month": grand_total, "ytd": ytd_grand_total},
        },
        "royalties": {
            "standard_rate": {
                "amount": standard_royalty["royalty"],
                "description": standard_royalty["description"],
                "tier": standard_royalty["tier_description"],
            },
            "reduced_rate": {
                "amount": reduced_royalty["royalty"],
                "description": reduced_royalty["description"],
                "excess_amount": reduced_royalty["excess_amount"],
            },
            "total": total_royalty_payable,
        },
        "fees": {
            "fixed_fee": fixed_fee,
            "national_accounts": national_accounts_fee,
            "national_brand_fund": {
                "amount": national_brand_fee,
                "capped": brand_fund["capped"],
                "remaining_cap": brand_fund["remaining_cap"],
            },
            "national_brand_reduced": national_brand_reduced_fee,
            "total": total_fees_payable,
        },
        "grand_total_payable": grand_total_payable,
    }


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def generate_payment_summary_excel(summary: dict, output_path: str) -> str:
    """
    Generate Payment Summary Excel file matching the SERVPRO format.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Summary Report"
    
    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    currency_font = Font(size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    
    light_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    row = 1
    
    # Header section
    ws.cell(row=row, column=1, value="Royalty Reporting").font = header_font
    row += 2
    
    # Franchise info
    month_name = datetime(2000, summary["period_month"], 1).strftime("%B") if summary["period_month"] else "N/A"
    
    ws.cell(row=row, column=1, value="Franchise:")
    ws.cell(row=row, column=2, value=f"{summary['franchise_number']} {summary['department_name']}")
    ws.cell(row=row, column=3, value="Owner:")
    ws.cell(row=row, column=4, value=summary.get("owner_name", ""))
    row += 1
    
    ws.cell(row=row, column=1, value="Month/Year:")
    ws.cell(row=row, column=2, value=f"{summary['period_month']:02d}/{summary['period_year']}")
    ws.cell(row=row, column=3, value="Date of Mailing:")
    ws.cell(row=row, column=4, value=datetime.now().strftime("%m/%d/%Y"))
    row += 2
    
    # Column headers
    headers = ["Category", "", "Revenue This Month", "Year To Date", "Royalty Payable"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Category data
    categories_config = [
        ("Water Restoration", "water"),
        ("Fire Restoration", "fire"),
        ("Mold/Bio Haz. Restoration", "mold_bio"),
        ("Other", "other"),
        ("Subcontract (Mitigation)", "subcontract"),
        ("Reconstruction (In-House & Subcontract)", "reconstruction"),
    ]
    
    for cat_name, cat_key in categories_config:
        cat_data = summary["categories"][cat_key]
        
        # Category header
        ws.cell(row=row, column=1, value=cat_name).font = section_font
        row += 1
        
        # Commercial
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="Commercial:")
        cell_rev = ws.cell(row=row, column=3, value=cat_data["commercial"]["this_month"])
        cell_rev.number_format = '#,##0.00'
        cell_rev.fill = yellow_fill
        cell_ytd = ws.cell(row=row, column=4, value=cat_data["commercial"]["ytd"])
        cell_ytd.number_format = '$#,##0.00'
        row += 1
        
        # Residential
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="Residential:")
        cell_rev = ws.cell(row=row, column=3, value=cat_data["residential"]["this_month"])
        cell_rev.number_format = '#,##0.00'
        cell_rev.fill = yellow_fill
        cell_ytd = ws.cell(row=row, column=4, value=cat_data["residential"]["ytd"])
        cell_ytd.number_format = '$#,##0.00'
        row += 1
    
    # Subtotals
    subtotals = summary["subtotals"]
    royalties = summary["royalties"]
    
    # Standard Rate Subtotal
    ws.cell(row=row, column=1, value="Subtotal: Standard Rate Services (Water, Fire, Mold/Bio Haz, Other)").font = section_font
    cell = ws.cell(row=row, column=3, value=subtotals["standard_rate"]["this_month"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=4, value=subtotals["standard_rate"]["ytd"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=5, value=royalties["standard_rate"]["amount"])
    cell.number_format = '$#,##0.00'
    row += 1
    
    # Reduced Rate Subtotal
    ws.cell(row=row, column=1, value="Subtotal: Reduced Rate Services (Subcontract, Reconstruction)").font = section_font
    cell = ws.cell(row=row, column=3, value=subtotals["reduced_rate"]["this_month"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=4, value=subtotals["reduced_rate"]["ytd"])
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=5, value=royalties["reduced_rate"]["amount"])
    cell.number_format = '$#,##0.00'
    row += 1
    
    # Total
    ws.cell(row=row, column=1, value="Total").font = section_font
    cell = ws.cell(row=row, column=3, value=subtotals["total"]["this_month"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    cell = ws.cell(row=row, column=4, value=subtotals["total"]["ytd"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    cell = ws.cell(row=row, column=5, value=royalties["total"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    row += 2
    
    # Fees section
    fees = summary["fees"]
    
    ws.cell(row=row, column=1, value="Fixed Fee")
    cell = ws.cell(row=row, column=5, value=fees["fixed_fee"])
    cell.number_format = '$#,##0.00'
    row += 1
    
    ws.cell(row=row, column=1, value="National Accounts Core/Regional Fee (0.5%)")
    cell = ws.cell(row=row, column=5, value=fees["national_accounts"])
    cell.number_format = '$#,##0.00'
    row += 1
    
    ws.cell(row=row, column=1, value="National Brand Fund Fee (2.5%)")
    cell = ws.cell(row=row, column=5, value=fees["national_brand_fund"]["amount"])
    cell.number_format = '$#,##0.00'
    row += 1
    
    ws.cell(row=row, column=1, value="National Brand Fund Fee - Reduced Rate Services (0.25%)")
    cell = ws.cell(row=row, column=5, value=fees["national_brand_reduced"])
    cell.number_format = '$#,##0.00'
    row += 2
    
    # Royalty calculation description
    std_royalty = royalties["standard_rate"]
    red_royalty = royalties["reduced_rate"]
    
    ws.cell(row=row, column=1, value="Royalty % And Fee Used").font = section_font
    ws.cell(row=row, column=4, value="Total Royalty, Fixed Fee,").font = section_font
    cell = ws.cell(row=row, column=5, value=summary["grand_total_payable"])
    cell.number_format = '$#,##0.00'
    cell.font = section_font
    row += 1
    
    ws.cell(row=row, column=1, value=std_royalty["description"])
    ws.cell(row=row, column=4, value="and Brand Fund Fee Payable").font = section_font
    row += 1
    
    ws.cell(row=row, column=1, value=f"plus ${fees['fixed_fee']:.2f} Fixed Fee")
    row += 1
    
    ws.cell(row=row, column=1, value=f"Reduced Rate Services {red_royalty['description']}")
    row += 2
    
    # Check fields
    ws.cell(row=row, column=4, value="Check Number")
    row += 1
    ws.cell(row=row, column=4, value="Check Amount")
    
    # Save
    wb.save(output_path)
    return output_path


def convert_excel_to_pdf(excel_path: str, pdf_path: str) -> bool:
    """
    Attempt to convert Excel to PDF using available methods.
    Returns True if successful, False otherwise.
    """
    # Try LibreOffice first
    try:
        import subprocess
        output_dir = os.path.dirname(pdf_path)
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', output_dir, excel_path],
            capture_output=True,
            timeout=60
        )
        if result.returncode == 0:
            # LibreOffice names the output based on input filename
            expected_pdf = os.path.join(output_dir, os.path.basename(excel_path).replace('.xlsx', '.pdf'))
            if os.path.exists(expected_pdf) and expected_pdf != pdf_path:
                os.rename(expected_pdf, pdf_path)
            return os.path.exists(pdf_path)
    except Exception as e:
        print(f"LibreOffice conversion failed: {e}")
    
    # Try win32com on Windows
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
        return os.path.exists(pdf_path)
    except Exception as e:
        print(f"win32com conversion failed: {e}")
    
    return False


# ============================================================================
# API ENDPOINTS
# ============================================================================

class PaymentSummaryRequest(BaseModel):
    realm_id: str
    department_id: str


@router.post("/generate")
async def generate_payment_summary(
    request: PaymentSummaryRequest,
    db: Session = Depends(get_db)
):
    """
    Generate Payment Summary (Royalty Report) for a specific franchise.
    """
    realm_id = request.realm_id
    department_id = request.department_id
    
    print(f"🚀 Starting Payment Summary generation for realm_id: {realm_id}, department_id: {department_id}")
    
    try:
        # 1. Get token from database
        token_record = db.query(QuickBooksToken).filter(
            QuickBooksToken.realm_id == realm_id
        ).first()
        
        if not token_record:
            raise HTTPException(status_code=404, detail=f"No QuickBooks token found for realm_id: {realm_id}")
        
        # 2. Refresh token if needed
        access_token = refresh_qbo_token(token_record, db)
        
        # 3. Get company info
        company_info = db.query(CompanyInfo).filter(
            CompanyInfo.realm_id == realm_id
        ).first()
        
        # 4. Get license mapping for department
        mapping = db.query(CompanyLicenseMapping).filter(
            CompanyLicenseMapping.realm_id == realm_id,
            CompanyLicenseMapping.qbo_department_id == department_id,
            CompanyLicenseMapping.is_active == True
        ).first()
        
        if not mapping:
            raise HTTPException(
                status_code=404, 
                detail=f"No license mapping found for department_id: {department_id}"
            )
        
        franchise_number = mapping.franchise_number
        department_name = mapping.qbo_department_name or "Unknown"
        
        print(f"📋 Franchise: {franchise_number}, Department: {department_name}")
        
        # 5. Query class IDs
        class_ids = query_class_ids(realm_id, access_token)
        
        # 6. Calculate period dates
        now = datetime.utcnow()
        last_month = now - relativedelta(months=1)
        period_start = datetime(last_month.year, last_month.month, 1).strftime("%Y-%m-%d")
        period_end = (datetime(last_month.year, last_month.month, 1) + relativedelta(months=1) - timedelta(days=1)).strftime("%Y-%m-%d")
        
        period_month = last_month.month
        period_year = last_month.year
        
        print(f"📅 Period: {period_start} to {period_end}")
        
        # 7. Fetch Last Month report
        last_month_data = fetch_class_sales_report(
            realm_id=realm_id,
            access_token=access_token,
            department_id=department_id,
            class_ids=class_ids,
            report_type="last_month"
        )
        
        # 8. Fetch YTD report
        ytd_data = fetch_class_sales_report(
            realm_id=realm_id,
            access_token=access_token,
            department_id=department_id,
            class_ids=class_ids,
            report_type="ytd"
        )
        
        # 9. Calculate payment summary
        summary = calculate_payment_summary(
            last_month_data=last_month_data,
            ytd_data=ytd_data,
            franchise_number=franchise_number,
            department_name=department_name,
            owner_name=company_info.company_name if company_info else "",
            period_month=period_month,
            period_year=period_year
        )
        
        # 10. Generate report name: Franchise # - mmyyyy Payment Summary
        report_name = f"{franchise_number} - {period_month:02d}{period_year} Payment Summary"
        
        # 11. Generate Excel file
        temp_dir = tempfile.mkdtemp()
        excel_path = os.path.join(temp_dir, f"{report_name}.xlsx")
        
        generate_payment_summary_excel(summary, excel_path)
        print(f"📊 Excel generated: {excel_path}")
        
        # 12. Try to generate PDF
        pdf_path = os.path.join(temp_dir, f"{report_name}.pdf")
        pdf_available = convert_excel_to_pdf(excel_path, pdf_path)
        
        if pdf_available:
            print(f"📄 PDF generated: {pdf_path}")
        else:
            print(f"ℹ️ PDF generation not available")
        
        # 13. Upload to Azure Storage
        storage = AzureStorageService()
        
        with open(excel_path, 'rb') as f:
            excel_bytes = f.read()
        
        excel_blob_url, excel_blob_name = storage.upload_file(
            file_data=excel_bytes,
            client_id=realm_id,
            license_id=franchise_number,
            file_name=report_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ext="xlsx"
        )
        print(f"☁️ Excel uploaded: {excel_blob_name}")
        
        pdf_blob_url = None
        pdf_blob_name = None
        if pdf_available and os.path.exists(pdf_path):
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            
            pdf_blob_url, pdf_blob_name = storage.upload_file(
                file_data=pdf_bytes,
                client_id=realm_id,
                license_id=franchise_number,
                file_name=report_name,
                content_type="application/pdf",
                ext="pdf"
            )
            print(f"☁️ PDF uploaded: {pdf_blob_name}")
        
        # 14. Generate SAS URLs (10 year expiry)
        excel_sas_url = storage.generate_sas_url(excel_blob_name, expiry_years=10)
        pdf_sas_url = storage.generate_sas_url(pdf_blob_name, expiry_years=10) if pdf_blob_name else None
        
        # 15. Save to database
        report_record = GeneratedReport(
            id=str(uuid.uuid4()),
            realm_id=realm_id,
            franchise_number=franchise_number,
            report_type="payment_summary",
            period_year=period_year,
            period_month=period_month,
            excel_blob_url=excel_blob_url,
            excel_blob_name=excel_blob_name,
            pdf_blob_url=pdf_blob_url,
            pdf_blob_name=pdf_blob_name,
            generated_at=datetime.utcnow(),
            generated_by="system"
        )
        db.add(report_record)
        db.commit()
        
        # 16. Cleanup temp files
        try:
            os.remove(excel_path)
            if pdf_available and os.path.exists(pdf_path):
                os.remove(pdf_path)
            os.rmdir(temp_dir)
        except Exception as e:
            print(f"⚠️ Temp cleanup warning: {e}")
        
        print(f"✅ Payment Summary generated successfully for {franchise_number}")
        
        return {
            "success": True,
            "report_id": report_record.id,
            "franchise_number": franchise_number,
            "department_name": department_name,
            "period": f"{period_month:02d}/{period_year}",
            "excel_download_url": excel_sas_url,
            "pdf_download_url": pdf_sas_url,
            "pdf_available": pdf_available,
            "summary": summary
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Payment Summary generation error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/generate-all/{realm_id}")
async def generate_all_payment_summaries(
    realm_id: str,
    db: Session = Depends(get_db)
):
    """
    Generate Payment Summary reports for all franchises of a company.
    """
    print(f"🚀 Starting bulk Payment Summary generation for realm_id: {realm_id}")
    
    # Get all active license mappings for this company
    mappings = db.query(CompanyLicenseMapping).filter(
        CompanyLicenseMapping.realm_id == realm_id,
        CompanyLicenseMapping.is_active == True
    ).all()
    
    if not mappings:
        raise HTTPException(
            status_code=404,
            detail=f"No active license mappings found for realm_id: {realm_id}"
        )
    
    results = []
    for mapping in mappings:
        try:
            request = PaymentSummaryRequest(
                realm_id=realm_id,
                department_id=mapping.qbo_department_id
            )
            result = await generate_payment_summary(request, db)
            results.append({
                "franchise_number": mapping.franchise_number,
                "department_name": mapping.qbo_department_name,
                "success": True,
                "report_id": result["report_id"],
                "excel_download_url": result["excel_download_url"],
                "pdf_download_url": result["pdf_download_url"],
            })
        except Exception as e:
            results.append({
                "franchise_number": mapping.franchise_number,
                "department_name": mapping.qbo_department_name,
                "success": False,
                "error": str(e)
            })
    
    success_count = sum(1 for r in results if r["success"])
    
    return {
        "realm_id": realm_id,
        "total": len(results),
        "success_count": success_count,
        "failed_count": len(results) - success_count,
        "results": results
    }


@router.get("/list/{realm_id}")
async def list_payment_summaries(
    realm_id: str,
    db: Session = Depends(get_db)
):
    """
    List all Payment Summary reports for a company.
    """
    reports = db.query(GeneratedReport).filter(
        GeneratedReport.realm_id == realm_id,
        GeneratedReport.report_type == "payment_summary"
    ).order_by(GeneratedReport.generated_at.desc()).all()
    
    storage = AzureStorageService()
    
    result = []
    for report in reports:
        excel_url = storage.generate_sas_url(report.excel_blob_name, expiry_years=10) if report.excel_blob_name else None
        pdf_url = storage.generate_sas_url(report.pdf_blob_name, expiry_years=10) if report.pdf_blob_name else None
        
        result.append({
            "id": report.id,
            "franchise_number": report.franchise_number,
            "period_year": report.period_year,
            "period_month": report.period_month,
            "generated_at": report.generated_at.isoformat() if report.generated_at else None,
            "excel_download_url": excel_url,
            "pdf_download_url": pdf_url,
        })
    
    return {
        "realm_id": realm_id,
        "count": len(result),
        "reports": result
    }


@router.get("/{report_id}")
async def get_payment_summary(
    report_id: str,
    db: Session = Depends(get_db)
):
    """
    Get a specific Payment Summary report.
    """
    report = db.query(GeneratedReport).filter(
        GeneratedReport.id == report_id,
        GeneratedReport.report_type == "payment_summary"
    ).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    storage = AzureStorageService()
    
    excel_url = storage.generate_sas_url(report.excel_blob_name, expiry_years=10) if report.excel_blob_name else None
    pdf_url = storage.generate_sas_url(report.pdf_blob_name, expiry_years=10) if report.pdf_blob_name else None
    
    return {
        "id": report.id,
        "realm_id": report.realm_id,
        "franchise_number": report.franchise_number,
        "period_year": report.period_year,
        "period_month": report.period_month,
        "generated_at": report.generated_at.isoformat() if report.generated_at else None,
        "excel_download_url": excel_url,
        "pdf_download_url": pdf_url,
    }

